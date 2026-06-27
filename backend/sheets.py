"""
sheets.py — Integrasi Google Sheets API
  • pull_master_data()          : Import 5 tab dari Sheets → Supabase
  • export_timetable_to_sheet() : Export jadwal → Sheets
Blueprint §2, §7
"""
import json
import logging
import io
import openpyxl
import gspread
from google.oauth2.service_account import Credentials

from backend.database import get_db_connection, db_fetchall, get_setting

logger = logging.getLogger(__name__)

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

DAYS_ORDER   = ["SENIN", "SELASA", "RABU", "KAMIS", "JUMAT", "SABTU"]
SHIFT_LIMITS = {
    "PAGI":  {"SENIN": 7, "SELASA": 7, "RABU": 7, "KAMIS": 7, "JUMAT": 6, "SABTU": 7},
    "SIANG": {"SENIN": 7, "SELASA": 7, "RABU": 7, "KAMIS": 7, "JUMAT": 6, "SABTU": 6},
}


# ─────────────────────────────────────────────
# Auth helpers
# ─────────────────────────────────────────────

def _gspread_client():
    creds_str = get_setting("credentials_json")
    if not creds_str:
        raise ValueError(
            "Credentials JSON belum dikonfigurasi. "
            "Isi di menu Pengaturan dashboard."
        )
    try:
        info  = json.loads(creds_str)
        creds = Credentials.from_service_account_info(info, scopes=SCOPES)
        return gspread.authorize(creds)
    except Exception as e:
        raise ValueError(f"Credentials JSON tidak valid: {e}")


def _open_spreadsheet(gc):
    sid = get_setting("spreadsheet_id")
    if not sid:
        raise ValueError(
            "Spreadsheet ID belum dikonfigurasi. "
            "Isi di menu Pengaturan dashboard."
        )
    try:
        return gc.open_by_key(sid)
    except Exception as e:
        raise ValueError(
            f"Gagal membuka spreadsheet [{sid}]. "
            f"Pastikan ID benar & Service Account sudah di-share sebagai Editor. ({e})"
        )


# ─────────────────────────────────────────────
# Import dari Google Sheets → Supabase
# ─────────────────────────────────────────────

def _import_data_to_db(rows_guru, rows_kelas, rows_mapel, rows_alokasi, rows_penugasan):
    # ── Validasi sequential dependency ─────────────────────────────
    if not rows_guru:
        raise ValueError("Tab 'master_guru' kosong. Isi data guru terlebih dahulu.")
    if not rows_kelas:
        raise ValueError("Tab 'master_kelas' kosong. Isi data kelas terlebih dahulu.")
    if not rows_mapel:
        raise ValueError("Tab 'master_mapel' kosong. Isi data mapel terlebih dahulu.")
    if not rows_alokasi:
        raise ValueError("Tab 'alokasi_kurikulum' kosong. Isi alokasi terlebih dahulu.")
    if not rows_penugasan:
        raise ValueError(
            "Tab 'penugasan_guru' kosong. "
            "Isi penugasan guru-mapel terlebih dahulu."
        )

    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        # Bersihkan semua data lama (CASCADE)
        cur.execute(
            "TRUNCATE TABLE timetable, class_subjects, teacher_subjects, "
            "teachers, classes, subjects RESTART IDENTITY CASCADE;"
        )

        # ── 1. Guru (Tab A) ─────────────────────────────────────────
        for r in rows_guru:
            nama = str(r.get("nama_guru", "")).strip()
            if not nama:
                continue
            hari_raw  = str(r.get("hari_tersedia", "")).strip()
            hari      = [h.strip().upper() for h in hari_raw.split(",") if h.strip()]

            pagi_raw  = str(r.get("hari_tersedia_pagi",  "")).strip()
            siang_raw = str(r.get("hari_tersedia_siang", "")).strip()
            hari_pagi  = [h.strip().upper() for h in pagi_raw.split(",")  if h.strip()] or hari
            hari_siang = [h.strip().upper() for h in siang_raw.split(",") if h.strip()] or hari

            shift_pagi  = str(r.get("shift_pagi",  "")).upper() in ("TRUE", "1", "YA", "YES")
            shift_siang = str(r.get("shift_siang", "")).upper() in ("TRUE", "1", "YA", "YES")

            # min_jp and max_jp values: NULL if empty
            try:
                min_jp = int(r.get("min_jp")) if str(r.get("min_jp", "")).strip() not in (None, "") else None
            except (ValueError, TypeError):
                min_jp = None
            try:
                max_jp = int(r.get("max_jp")) if str(r.get("max_jp", "")).strip() not in (None, "") else None
            except (ValueError, TypeError):
                max_jp = None

            allowed_jp_pagi_raw = r.get("allowed_jp_pagi")
            allowed_jp_pagi = None
            if allowed_jp_pagi_raw:
                try:
                    if isinstance(allowed_jp_pagi_raw, dict):
                        allowed_jp_pagi = json.dumps(allowed_jp_pagi_raw)
                    elif isinstance(allowed_jp_pagi_raw, str) and allowed_jp_pagi_raw.strip():
                        allowed_jp_pagi = json.dumps(json.loads(allowed_jp_pagi_raw))
                except Exception:
                    pass

            allowed_jp_siang_raw = r.get("allowed_jp_siang")
            allowed_jp_siang = None
            if allowed_jp_siang_raw:
                try:
                    if isinstance(allowed_jp_siang_raw, dict):
                        allowed_jp_siang = json.dumps(allowed_jp_siang_raw)
                    elif isinstance(allowed_jp_siang_raw, str) and allowed_jp_siang_raw.strip():
                        allowed_jp_siang = json.dumps(json.loads(allowed_jp_siang_raw))
                except Exception:
                    pass

            no_wa = str(r.get("no_wa", "")).strip() if r.get("no_wa") not in (None, "") else None

            cur.execute("""
                INSERT INTO teachers
                    (nama_guru, kode_guru, hari_tersedia,
                     shift_pagi, shift_siang,
                     hari_tersedia_pagi, hari_tersedia_siang, min_jp, max_jp,
                     allowed_jp_pagi, allowed_jp_siang, no_wa)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                nama,
                int(r.get("kode_guru")),
                json.dumps(hari),
                shift_pagi, shift_siang,
                json.dumps(hari_pagi),
                json.dumps(hari_siang),
                min_jp, max_jp,
                allowed_jp_pagi, allowed_jp_siang,
                no_wa
            ))

        # ── 2. Kelas (Tab B) ────────────────────────────────────────
        for r in rows_kelas:
            nama_kelas = str(r.get("nama_kelas", "")).strip()
            if not nama_kelas:
                continue
            shift_op = str(r.get("shift_operasional", "")).strip().upper()

            # Ambil tingkat & jurusan dari kolom sheet, fallback auto-parse
            tingkat_val = str(r.get("tingkat", "")).strip().upper() or None
            jurusan_val = str(r.get("jurusan", "")).strip().upper() or None
            if not tingkat_val or not jurusan_val:
                parts = nama_kelas.split()
                if len(parts) >= 2:
                    tk = parts[0].upper()
                    if tk in ("X", "XI", "XII"):
                        tingkat_val = tingkat_val or tk
                        jurusan_val = jurusan_val or parts[1].upper()

            cur.execute(
                "INSERT INTO classes (nama_kelas, shift_operasional, tingkat, jurusan) "
                "VALUES (%s,%s,%s,%s)",
                (nama_kelas, shift_op, tingkat_val, jurusan_val)
            )

        # ── 3. Mapel (Tab C) ────────────────────────────────────────
        for r in rows_mapel:
            nama_mapel = str(r.get("nama_mapel", "")).strip()
            if not nama_mapel:
                continue
            cur.execute(
                "INSERT INTO subjects (nama_mapel, kategori_mapel, tingkat, jurusan) "
                "VALUES (%s,%s,%s,%s)",
                (
                    nama_mapel,
                    str(r.get("kategori_mapel", "UMUM")).strip().upper(),
                    str(r.get("tingkat", "")).strip().upper() or None,
                    str(r.get("jurusan",  "")).strip().upper() or None,
                )
            )

        conn.commit()

        # ── Build lookup maps ───────────────────────────────────────
        class_map = {
            r["nama_kelas"]: r["id_kelas"]
            for r in db_fetchall(conn, "SELECT id_kelas, nama_kelas FROM classes")
        }
        subj_map = {
            r["nama_mapel"]: r["id_mapel"]
            for r in db_fetchall(conn, "SELECT id_mapel, nama_mapel FROM subjects")
        }
        teacher_name_map = {
            r["nama_guru"].strip().upper(): r["id_guru"]
            for r in db_fetchall(conn, "SELECT id_guru, nama_guru FROM teachers")
        }

        # ── 4. Alokasi Kurikulum (Tab D) — TANPA guru ───────────────
        alokasi_errors = []
        for r in rows_alokasi:
            nama_kelas = str(r.get("nama_kelas", "")).strip()
            nama_mapel = str(r.get("nama_mapel", "")).strip()
            durasi     = int(r.get("durasi_jp", 0) or 0)

            id_kelas = class_map.get(nama_kelas)
            id_mapel = subj_map.get(nama_mapel)

            if not id_kelas:
                alokasi_errors.append(f"Kelas [{nama_kelas}] tidak ditemukan di master_kelas.")
                continue
            if not id_mapel:
                alokasi_errors.append(f"Mapel [{nama_mapel}] tidak ditemukan di master_mapel.")
                continue
            if durasi <= 0:
                alokasi_errors.append(f"Durasi JP tidak valid untuk [{nama_kelas}] - [{nama_mapel}].")
                continue

            cur.execute(
                "INSERT INTO class_subjects (id_kelas, id_mapel, durasi_jp) "
                "VALUES (%s,%s,%s) ON CONFLICT DO NOTHING",
                (id_kelas, id_mapel, durasi)
            )

        # ── 5. Penugasan Guru (Tab E) ────────────────────────────────
        penugasan_errors = []
        for r in rows_penugasan:
            nama_guru  = str(r.get("nama_guru",  "")).strip()
            nama_mapel = str(r.get("nama_mapel", "")).strip()

            id_guru  = teacher_name_map.get(nama_guru.upper())
            id_mapel = subj_map.get(nama_mapel)

            if not id_guru:
                # Coba partial match
                for k, v in teacher_name_map.items():
                    if nama_guru.upper() in k:
                        id_guru = v
                        break
            if not id_guru:
                penugasan_errors.append(f"Guru [{nama_guru}] tidak ditemukan di master_guru.")
                continue
            if not id_mapel:
                penugasan_errors.append(f"Mapel [{nama_mapel}] tidak ditemukan di master_mapel.")
                continue

            cur.execute(
                "INSERT INTO teacher_subjects (id_guru, id_mapel) "
                "VALUES (%s,%s) ON CONFLICT DO NOTHING",
                (id_guru, id_mapel)
            )

        conn.commit()

        # ── Coverage check setelah sync ─────────────────────────────
        all_teachers = db_fetchall(conn, "SELECT * FROM teachers")
        all_classes  = db_fetchall(conn, "SELECT * FROM classes")
        for t in all_teachers:
            t["hari_tersedia_pagi"]  = json.loads(t["hari_tersedia_pagi"]  or "[]")
            t["hari_tersedia_siang"] = json.loads(t["hari_tersedia_siang"] or "[]")
            t["shift_pagi"]  = bool(t["shift_pagi"])
            t["shift_siang"] = bool(t["shift_siang"])

        coverage_warnings = []
        for shift in ["PAGI", "SIANG"]:
            n_kelas = sum(1 for c in all_classes if c["shift_operasional"] == shift)
            if n_kelas == 0:
                continue
            for day in DAYS_ORDER:
                if shift == "PAGI":
                    n_guru = sum(
                        1 for t in all_teachers
                        if t["shift_pagi"] and day in t["hari_tersedia_pagi"]
                    )
                else:
                    n_guru = sum(
                        1 for t in all_teachers
                        if t["shift_siang"] and day in t["hari_tersedia_siang"]
                    )
                if n_guru < n_kelas:
                    coverage_warnings.append(
                        f"⚠️ Coverage — Shift {shift}, {day}: "
                        f"{n_kelas} kelas aktif, {n_guru} guru tersedia "
                        f"(kekurangan {n_kelas - n_guru})"
                    )

        all_errors = alokasi_errors + penugasan_errors

        logger.info("Sinkronisasi data master dari Google Sheets selesai.")
        return {
            "status":  "SUCCESS" if not all_errors else "PARTIAL",
            "message": "Sinkronisasi selesai!",
            "errors":  all_errors,
            "coverage_warnings": coverage_warnings,
            "stats": {
                "guru":      len(rows_guru),
                "kelas":     len(rows_kelas),
                "mapel":     len(rows_mapel),
                "alokasi":   len(rows_alokasi),
                "penugasan": len(rows_penugasan),
            },
        }

    except Exception:
        conn.rollback()
        logger.exception("Sync gagal")
        raise
    finally:
        cur.close()
        conn.close()

def pull_master_data():
    """
    Membaca 5 tab Google Sheets secara sequential dan menimpa database.
    Tab: master_guru → master_kelas → master_mapel →
         alokasi_kurikulum → penugasan_guru
    Blueprint §2.
    """
    gc = _gspread_client()
    sh = _open_spreadsheet(gc)

    def _read(tab_name):
        try:
            return sh.worksheet(tab_name).get_all_records()
        except Exception as e:
            raise ValueError(f"Gagal membaca sheet '{tab_name}': {e}")

    rows_guru       = _read("master_guru")
    rows_kelas      = _read("master_kelas")
    rows_mapel      = _read("master_mapel")
    rows_alokasi    = _read("alokasi_kurikulum")
    rows_penugasan  = _read("penugasan_guru")

    return _import_data_to_db(rows_guru, rows_kelas, rows_mapel, rows_alokasi, rows_penugasan)

def pull_excel_data(file_bytes):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Gagal membaca file Excel: {e}")

    def _read(tab_name):
        if tab_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{tab_name}' tidak ditemukan di file Excel.")
        ws = wb[tab_name]
        rows = list(ws.values)
        if not rows or len(rows) < 2:
            return []
        headers = [str(h).strip() if h else "" for h in rows[0]]
        result = []
        for row in rows[1:]:
            # Abaikan row jika semua cell kosong
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            d = dict(zip(headers, row))
            result.append(d)
        return result

    rows_guru       = _read("master_guru")
    rows_kelas      = _read("master_kelas")
    rows_mapel      = _read("master_mapel")
    rows_alokasi    = _read("alokasi_kurikulum")
    rows_penugasan  = _read("penugasan_guru")

    return _import_data_to_db(rows_guru, rows_kelas, rows_mapel, rows_alokasi, rows_penugasan)


# ─────────────────────────────────────────────
# Export jadwal → Google Sheets
# ─────────────────────────────────────────────

def export_timetable_to_sheet():
    """
    Menulis hasil jadwal ke tab 'hasil_jadwal' dengan format
    kolom berdampingan per kelas (2 kolom: Mapel | Kode Guru).
    Blueprint §7B, §7C.
    """
    gc = _gspread_client()
    sh = _open_spreadsheet(gc)

    try:
        ws = sh.worksheet("hasil_jadwal")
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title="hasil_jadwal", rows=200, cols=150)

    conn = get_db_connection()
    try:
        classes  = db_fetchall(conn, "SELECT * FROM classes ORDER BY nama_kelas")
        teachers = {
            r["id_guru"]: r
            for r in db_fetchall(conn, "SELECT id_guru, kode_guru, nama_guru FROM teachers")
        }

        # Ambil timetable dengan JOIN via id_class_subject
        tt_rows = db_fetchall(conn, """
            SELECT
                t.hari, t.jam_ke, t.id_guru, t.is_fallback,
                cs.id_kelas, cs.id_mapel,
                s.nama_mapel
            FROM timetable t
            JOIN class_subjects cs ON t.id_class_subject = cs.id_class_subject
            JOIN subjects s ON cs.id_mapel = s.id_mapel
        """)
    finally:
        conn.close()

    if not classes:
        raise ValueError("Tidak ada data kelas. Tarik data master terlebih dahulu.")

    # Index: (id_kelas, hari, jam_ke) → row
    tt = {(r["id_kelas"], r["hari"], r["jam_ke"]): r for r in tt_rows}

    # Build grid 2D
    h1 = ["HARI", "JAM KE"]
    h2 = ["",     ""]
    for c in classes:
        h1 += [c["nama_kelas"], c["nama_kelas"]]
        h2 += ["Mata Pelajaran", "Kode Guru"]

    grid = [h1, h2]

    for day in DAYS_ORDER:
        for jp in range(1, 8):
            row = [day, jp]
            for c in classes:
                cid    = c["id_kelas"]
                shift  = c["shift_operasional"]
                max_jp = SHIFT_LIMITS[shift].get(day, 0)
                
                # Check for Monday Pagi Upacara
                if day == "SENIN" and jp == 1 and shift == "PAGI":
                    row += ["UPACARA BENDERA", ""]
                elif jp > max_jp:
                    row += ["-", "-"]
                else:
                    entry = tt.get((cid, day, jp))
                    if entry:
                        mapel = entry.get("nama_mapel", "")
                        if entry.get("is_fallback"):
                            mapel += " *"
                        t_info = teachers.get(entry["id_guru"])
                        kode   = t_info["kode_guru"] if t_info else ""
                        row += [mapel, kode]
                    else:
                        row += ["KOSONG", ""]
            grid.append(row)
            
            # Insert ISTIRAHAT row after JP 4 for all days
            if jp == 4:
                break_row = [day, "ISTIRAHAT"]
                for c in classes:
                    break_row += ["ISTIRAHAT", ""]
                grid.append(break_row)

    # Kirim sekaligus (Blueprint §7C — mencegah 429)
    ws.clear()
    ws.update(grid, "A1")

    return {
        "status":  "SUCCESS",
        "message": "Jadwal berhasil diekspor ke sheet 'hasil_jadwal'!",
    }


# ─────────────────────────────────────────────
# Bulk Upload Excel Mandiri
# ─────────────────────────────────────────────

def import_teachers_from_excel(file_bytes):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Gagal membaca file Excel: {e}")
    
    sheet_name = "master_guru" if "master_guru" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.values)
    if not rows or len(rows) < 2:
        raise ValueError(f"Sheet '{sheet_name}' kosong atau tidak memiliki data.")
    
    headers = [str(h).strip() if h else "" for h in rows[0]]
    
    conn = get_db_connection()
    cur = conn.cursor()
    success_count = 0
    update_count = 0
    errors = []
    
    try:
        for idx, row in enumerate(rows[1:], start=2):
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            r = dict(zip(headers, row))
            nama = str(r.get("nama_guru", "")).strip()
            kode_str = str(r.get("kode_guru", "")).strip()
            
            if not nama or not kode_str:
                errors.append(f"Baris {idx}: nama_guru atau kode_guru kosong.")
                continue
            
            try:
                kode = int(float(kode_str))
            except (ValueError, TypeError):
                errors.append(f"Baris {idx}: kode_guru harus berupa angka (ditemukan: '{kode_str}').")
                continue
                
            hari_raw  = str(r.get("hari_tersedia", "")).strip()
            hari      = [h.strip().upper() for h in hari_raw.split(",") if h.strip()]

            pagi_raw  = str(r.get("hari_tersedia_pagi",  "")).strip()
            siang_raw = str(r.get("hari_tersedia_siang", "")).strip()
            hari_pagi  = [h.strip().upper() for h in pagi_raw.split(",")  if h.strip()] or hari
            hari_siang = [h.strip().upper() for h in siang_raw.split(",") if h.strip()] or hari

            shift_pagi  = str(r.get("shift_pagi",  "")).upper() in ("TRUE", "1", "YA", "YES")
            shift_siang = str(r.get("shift_siang", "")).upper() in ("TRUE", "1", "YA", "YES")

            try:
                min_jp = int(float(r.get("min_jp"))) if r.get("min_jp") not in (None, "") else 2
            except (ValueError, TypeError):
                min_jp = 2
            try:
                max_jp = int(float(r.get("max_jp"))) if r.get("max_jp") not in (None, "") else 60
            except (ValueError, TypeError):
                max_jp = 60

            allowed_jp_pagi_raw = r.get("allowed_jp_pagi")
            allowed_jp_pagi = None
            if allowed_jp_pagi_raw:
                try:
                    if isinstance(allowed_jp_pagi_raw, dict):
                        allowed_jp_pagi = json.dumps(allowed_jp_pagi_raw)
                    elif isinstance(allowed_jp_pagi_raw, str) and allowed_jp_pagi_raw.strip():
                        allowed_jp_pagi = json.dumps(json.loads(allowed_jp_pagi_raw))
                except Exception:
                    pass

            allowed_jp_siang_raw = r.get("allowed_jp_siang")
            allowed_jp_siang = None
            if allowed_jp_siang_raw:
                try:
                    if isinstance(allowed_jp_siang_raw, dict):
                        allowed_jp_siang = json.dumps(allowed_jp_siang_raw)
                    elif isinstance(allowed_jp_siang_raw, str) and allowed_jp_siang_raw.strip():
                        allowed_jp_siang = json.dumps(json.loads(allowed_jp_siang_raw))
                except Exception:
                    pass

            no_wa = str(r.get("no_wa", "")).strip() if r.get("no_wa") not in (None, "") else None

            # Check if exists to determine insert vs update
            cur.execute("SELECT id_guru FROM teachers WHERE kode_guru = %s", (kode,))
            exists = cur.fetchone()
            
            cur.execute("""
                INSERT INTO teachers
                    (nama_guru, kode_guru, hari_tersedia,
                     shift_pagi, shift_siang,
                     hari_tersedia_pagi, hari_tersedia_siang, min_jp, max_jp,
                     allowed_jp_pagi, allowed_jp_siang, no_wa)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (kode_guru) DO UPDATE SET
                    nama_guru = EXCLUDED.nama_guru,
                    hari_tersedia = EXCLUDED.hari_tersedia,
                    shift_pagi = EXCLUDED.shift_pagi,
                    shift_siang = EXCLUDED.shift_siang,
                    hari_tersedia_pagi = EXCLUDED.hari_tersedia_pagi,
                    hari_tersedia_siang = EXCLUDED.hari_tersedia_siang,
                    min_jp = EXCLUDED.min_jp,
                    max_jp = EXCLUDED.max_jp,
                    allowed_jp_pagi = COALESCE(EXCLUDED.allowed_jp_pagi, teachers.allowed_jp_pagi),
                    allowed_jp_siang = COALESCE(EXCLUDED.allowed_jp_siang, teachers.allowed_jp_siang),
                    no_wa = EXCLUDED.no_wa
            """, (
                nama, kode, json.dumps(hari),
                shift_pagi, shift_siang,
                json.dumps(hari_pagi), json.dumps(hari_siang),
                min_jp, max_jp,
                allowed_jp_pagi, allowed_jp_siang,
                no_wa
            ))
            if exists:
                update_count += 1
            else:
                success_count += 1
                
        conn.commit()
        return {
            "status": "SUCCESS" if not errors else "PARTIAL",
            "added": success_count,
            "updated": update_count,
            "errors": errors
        }
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        cur.close()
        conn.close()

def import_subjects_from_excel(file_bytes):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Gagal membaca file Excel: {e}")
    
    sheet_name = "master_mapel" if "master_mapel" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.values)
    if not rows or len(rows) < 2:
        raise ValueError(f"Sheet '{sheet_name}' kosong atau tidak memiliki data.")
    
    headers = [str(h).strip() if h else "" for h in rows[0]]
    
    conn = get_db_connection()
    cur = conn.cursor()
    success_count = 0
    errors = []
    
    try:
        for idx, row in enumerate(rows[1:], start=2):
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            r = dict(zip(headers, row))
            nama = str(r.get("nama_mapel", "")).strip()
            kategori = str(r.get("kategori_mapel", "UMUM")).strip().upper()
            tingkat = str(r.get("tingkat", "")).strip().upper() or None
            jurusan = str(r.get("jurusan", "")).strip().upper() or None
            
            if not nama:
                errors.append(f"Baris {idx}: nama_mapel kosong.")
                continue
            
            # Check for duplicate
            cur.execute("""
                SELECT id_mapel FROM subjects 
                WHERE UPPER(nama_mapel) = UPPER(%s) 
                  AND (UPPER(tingkat) = UPPER(%s) OR (tingkat IS NULL AND %s IS NULL))
                  AND (UPPER(jurusan) = UPPER(%s) OR (jurusan IS NULL AND %s IS NULL))
            """, (nama, tingkat, tingkat, jurusan, jurusan))
            
            if cur.fetchone():
                continue
                
            cur.execute("""
                INSERT INTO subjects (nama_mapel, kategori_mapel, tingkat, jurusan)
                VALUES (%s,%s,%s,%s)
            """, (nama, kategori, tingkat, jurusan))
            success_count += 1
            
        conn.commit()
        return {
            "status": "SUCCESS" if not errors else "PARTIAL",
            "added": success_count,
            "errors": errors
        }
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        cur.close()
        conn.close()

def import_teacher_subjects_from_excel(file_bytes):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Gagal membaca file Excel: {e}")
    
    sheet_name = "penugasan_guru" if "penugasan_guru" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.values)
    if not rows or len(rows) < 2:
        raise ValueError(f"Sheet '{sheet_name}' kosong atau tidak memiliki data.")
    
    headers = [str(h).strip() if h else "" for h in rows[0]]
    
    conn = get_db_connection()
    cur = conn.cursor()
    success_count = 0
    errors = []
    
    try:
        for idx, row in enumerate(rows[1:], start=2):
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            r = dict(zip(headers, row))
            nama_guru = str(r.get("nama_guru", "")).strip()
            kode_guru_str = str(r.get("kode_guru", "")).strip()
            nama_mapel = str(r.get("nama_mapel", "")).strip()
            
            if not nama_mapel:
                errors.append(f"Baris {idx}: nama_mapel kosong.")
                continue
            if not nama_guru and not kode_guru_str:
                errors.append(f"Baris {idx}: nama_guru dan kode_guru kosong.")
                continue
                
            # Find teacher id_guru
            id_guru = None
            if kode_guru_str:
                try:
                    kode_guru = int(float(kode_guru_str))
                    cur.execute("SELECT id_guru FROM teachers WHERE kode_guru = %s", (kode_guru,))
                    res = cur.fetchone()
                    if res:
                        id_guru = res["id_guru"]
                except ValueError:
                    pass
            
            if not id_guru and nama_guru:
                cur.execute("SELECT id_guru FROM teachers WHERE UPPER(nama_guru) = UPPER(%s)", (nama_guru,))
                res = cur.fetchone()
                if res:
                    id_guru = res["id_guru"]
            
            if not id_guru:
                target_identity = f"Kode: {kode_guru_str}" if kode_guru_str else f"Nama: {nama_guru}"
                errors.append(f"Baris {idx}: Guru [{target_identity}] tidak ditemukan di database.")
                continue
                
            # Find subject id_mapel
            cur.execute("SELECT id_mapel FROM subjects WHERE UPPER(nama_mapel) = UPPER(%s)", (nama_mapel,))
            subjects_res = cur.fetchall()
            if not subjects_res:
                errors.append(f"Baris {idx}: Mata pelajaran [{nama_mapel}] tidak ditemukan di database.")
                continue
                
            inserted_any = False
            for s_res in subjects_res:
                id_mapel = s_res["id_mapel"]
                try:
                    cur.execute("""
                        INSERT INTO teacher_subjects (id_guru, id_mapel)
                        VALUES (%s,%s)
                        ON CONFLICT (id_guru, id_mapel) DO NOTHING
                    """, (id_guru, id_mapel))
                    inserted_any = True
                except Exception:
                    pass
            
            if inserted_any:
                success_count += 1
                
        conn.commit()
        return {
            "status": "SUCCESS" if not errors else "PARTIAL",
            "added": success_count,
            "errors": errors
        }
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        cur.close()
        conn.close()
