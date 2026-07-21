"""
main.py - FastAPI application entry point
SITAB: Automatic Timetable Generator (Dual-Shift SMK)
Database: Supabase (PostgreSQL via psycopg2)
Schema: class_subjects tanpa id_guru; timetable via id_class_subject
"""
import os
os.environ["OPENBLAS_NUM_THREADS"] = "1"
import io
import json
import logging
from typing import List, Optional

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import pymysql
from fastapi import FastAPI, HTTPException, UploadFile, File, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import Workbook

from backend.database import (
    init_db, get_db_connection, db_fetchall, db_fetchone,
    get_setting, save_setting, clear_master_data, active_branch, set_thread_branch,
    get_all_lms_endpoints, create_lms_endpoint, update_lms_endpoint,
    delete_lms_endpoint, set_active_lms_endpoint, get_active_lms_endpoint,
    get_current_db_name, get_branch_name,
)
from backend.models import (
    Teacher, TeacherCreate, TeacherAvailabilityBatchItem,
    Class,   ClassCreate,
    Subject, SubjectCreate,
    ClassSubject, ClassSubjectCreate,
    TeacherSubject, TeacherSubjectCreate,
    SettingsUpdate, AllocationUpdate, AllocationCopy,
    TimeSlotItem, TimeSlotBulkSave, TimeSlotCopy,
)
from backend.solver import generate_timetable, _diagnose_infeasibility, _preflight
from backend.sheets import (
    pull_master_data, export_timetable_to_sheet, pull_excel_data,
    import_teachers_from_excel, import_subjects_from_excel, import_teacher_subjects_from_excel,
)
from backend.excel_generator import generate_excel_timetable


logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("main")

init_db()

app = FastAPI(title="SITAB - Automatic Timetable Generator")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_credentials=True,
    allow_methods=["*"], allow_headers=["*"],
)

@app.middleware("http")
async def db_schema_middleware(request, call_next):
    return await call_next(request)

DAYS_ORDER   = ["SENIN", "SELASA", "RABU", "KAMIS", "JUMAT", "SABTU"]
SHIFT_LIMITS = {
    "PAGI":  {"SENIN": 7, "SELASA": 7, "RABU": 7, "KAMIS": 7, "JUMAT": 6, "SABTU": 7},
    "SIANG": {"SENIN": 7, "SELASA": 7, "RABU": 7, "KAMIS": 7, "JUMAT": 6, "SABTU": 6},
}


# ═══════════════════════════════════════════════
#  Settings
# ═══════════════════════════════════════════════

@app.get("/api/info")
def get_app_info():
    return {
        "db_name": get_current_db_name(),
        "branch_name": get_branch_name()
    }

@app.get("/api/settings")
def get_system_settings():
    return {
        "spreadsheet_id":   get_setting("spreadsheet_id", ""),
        "credentials_json": get_setting("credentials_json", ""),
        "lms_api_url":      get_setting("lms_api_url", ""),
        "lms_api_key":      get_setting("lms_api_key", ""),
    }

@app.post("/api/settings")
def update_system_settings(body: SettingsUpdate):
    if body.credentials_json:
        try:
            json.loads(body.credentials_json)
        except json.JSONDecodeError:
            raise HTTPException(400, "Format Credentials JSON tidak valid!")
    save_setting("spreadsheet_id",   body.spreadsheet_id)
    save_setting("credentials_json", body.credentials_json)
    save_setting("lms_api_url",      body.lms_api_url or "")
    save_setting("lms_api_key",      body.lms_api_key or "")
    return {"status": "SUCCESS", "message": "Pengaturan berhasil disimpan!"}


# ═══════════════════════════════════════════════
#  Time Slot Settings (Waktu Jam Pelajaran per Hari & Shift)
# ═══════════════════════════════════════════════



# ═══════════════════════════════════════════════
#  JP Limit Settings (Batas JP per Guru per Kelas per Hari)
# ═══════════════════════════════════════════════

from pydantic import BaseModel as _BM

@app.get("/api/settings/jp-limits")
def get_jp_limits():
    """Ambil batas JP per guru per kelas per hari (ideal & darurat) serta aturan multi-mapel."""
    try:
        ideal   = int(get_setting("max_jp_ideal",   "3"))
        darurat = int(get_setting("max_jp_darurat", "4"))
        split_multi = get_setting("split_multi_subject_teacher_days", "true").lower() == "true"
        multi_threshold = int(get_setting("multi_subject_jp_threshold", "4"))
    except (ValueError, TypeError):
        ideal, darurat, split_multi, multi_threshold = 3, 4, True, 4
    return {
        "max_jp_ideal": ideal,
        "max_jp_darurat": darurat,
        "split_multi_subject": split_multi,
        "multi_subject_jp_threshold": multi_threshold,
    }


class JpLimitsUpdate(_BM):
    max_jp_ideal:   int
    max_jp_darurat: int
    split_multi_subject: bool = True
    multi_subject_jp_threshold: int = 4


@app.post("/api/settings/jp-limits")
def update_jp_limits(body: JpLimitsUpdate):
    """Simpan batas JP per guru per kelas per hari & aturan pemisahan hari guru multi-mapel."""
    if body.max_jp_ideal < 1 or body.max_jp_darurat < 1:
        raise HTTPException(400, "Nilai batas JP minimal 1.")
    if body.max_jp_ideal > body.max_jp_darurat:
        raise HTTPException(400, "Batas ideal tidak boleh melebihi batas darurat.")
    if body.multi_subject_jp_threshold < 1:
        raise HTTPException(400, "Batas threshold jam multi-mapel minimal 1 JP.")

    save_setting("max_jp_ideal",   str(body.max_jp_ideal))
    save_setting("max_jp_darurat", str(body.max_jp_darurat))
    save_setting("split_multi_subject_teacher_days", "true" if body.split_multi_subject else "false")
    save_setting("multi_subject_jp_threshold", str(body.multi_subject_jp_threshold))

    return {
        "status": "SUCCESS",
        "message": f"Pengaturan disimpan: ideal={body.max_jp_ideal} JP, darurat={body.max_jp_darurat} JP, pisah_hari={body.split_multi_subject}, threshold={body.multi_subject_jp_threshold} JP.",
        "max_jp_ideal":   body.max_jp_ideal,
        "max_jp_darurat": body.max_jp_darurat,
        "split_multi_subject": body.split_multi_subject,
        "multi_subject_jp_threshold": body.multi_subject_jp_threshold,
    }


# ═══════════════════════════════════════════════
#  LMS Endpoint Management (daftar endpoint, pilih aktif)
# ═══════════════════════════════════════════════


class LmsEndpointCreate(_BM):
    nama_label:   str
    endpoint_url: str
    bearer_token: str
    keterangan:   str = ""

class LmsEndpointUpdate(_BM):
    nama_label:   str
    endpoint_url: str
    bearer_token: str = ""   # kosong = tidak ganti token
    keterangan:   str = ""


@app.get("/api/lms-endpoints")
def list_lms_endpoints():
    rows = get_all_lms_endpoints()
    # Sembunyikan token asli, hanya tampilkan 6 karakter pertama
    result = []
    for r in rows:
        d = dict(r)
        token = d.get("bearer_token", "")
        d["bearer_token_preview"] = token[:6] + "*" * max(0, len(token) - 6) if token else ""
        d["bearer_token"] = ""  # jangan expose ke frontend
        result.append(d)
    return result


@app.post("/api/lms-endpoints")
def add_lms_endpoint(body: LmsEndpointCreate):
    new_ep = create_lms_endpoint(
        nama_label=body.nama_label,
        endpoint_url=body.endpoint_url,
        bearer_token=body.bearer_token,
        keterangan=body.keterangan,
    )
    return {"status": "SUCCESS", "message": f"Endpoint '{body.nama_label}' berhasil ditambahkan.", "data": new_ep}


@app.put("/api/lms-endpoints/{endpoint_id}")
def edit_lms_endpoint(endpoint_id: int, body: LmsEndpointUpdate):
    ok = update_lms_endpoint(
        endpoint_id=endpoint_id,
        nama_label=body.nama_label,
        endpoint_url=body.endpoint_url,
        bearer_token=body.bearer_token or None,
        keterangan=body.keterangan,
    )
    if not ok:
        raise HTTPException(404, "Endpoint tidak ditemukan.")
    return {"status": "SUCCESS", "message": "Endpoint berhasil diperbarui."}


@app.delete("/api/lms-endpoints/{endpoint_id}")
def remove_lms_endpoint(endpoint_id: int):
    ok = delete_lms_endpoint(endpoint_id)
    if not ok:
        raise HTTPException(404, "Endpoint tidak ditemukan.")
    return {"status": "SUCCESS", "message": "Endpoint berhasil dihapus."}


@app.post("/api/lms-endpoints/{endpoint_id}/set-active")
def activate_lms_endpoint(endpoint_id: int):
    ok = set_active_lms_endpoint(endpoint_id)
    if not ok:
        raise HTTPException(404, "Endpoint tidak ditemukan.")
    return {"status": "SUCCESS", "message": "Endpoint aktif berhasil diperbarui."}


# ═══════════════════════════════════════════════
#  Clear all data
# ═══════════════════════════════════════════════

@app.post("/api/clear")
def clear_all_data():
    try:
        clear_master_data()
        return {"status": "SUCCESS", "message": "Semua data master & jadwal berhasil dihapus!"}
    except Exception as e:
        raise HTTPException(500, str(e))


# ═══════════════════════════════════════════════
#  Template Excel (5 sheets)
# ═══════════════════════════════════════════════

@app.get("/api/template")
def download_template():
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "master_guru"
    ws1.append(["nama_guru", "kode_guru", "hari_tersedia", "shift_pagi", "shift_siang",
                 "hari_tersedia_pagi", "hari_tersedia_siang", "min_jp", "max_jp", "no_wa"])
    ws1.append(["Budi Santoso, S.Pd",  101, "SENIN,SELASA,RABU,KAMIS",               "YA", "TIDAK", "SENIN,SELASA,RABU,KAMIS", "", "", "", "081234567890"])
    ws1.append(["Siti Aminah, S.Pd",   102, "SENIN,RABU,KAMIS,JUMAT,SABTU",           "YA", "YA",    "SENIN,RABU,KAMIS", "JUMAT,SABTU", "", "", "089876543210"])
    ws1.append(["Eko Prasetyo, S.Kom", 103, "SENIN,SELASA,RABU,KAMIS,JUMAT,SABTU",   "YA", "YA",    "SENIN,SELASA,RABU,KAMIS,JUMAT,SABTU", "SENIN,SELASA,RABU,KAMIS,JUMAT,SABTU", "", "", "085544332211"])

    ws2 = wb.create_sheet("master_kelas")
    ws2.append(["nama_kelas", "shift_operasional", "tingkat", "jurusan"])
    ws2.append(["X TKJ 1",  "PAGI",  "X",  "TKJ"])
    ws2.append(["XI AKL 1", "SIANG", "XI", "AKL"])

    ws3 = wb.create_sheet("master_mapel")
    ws3.append(["nama_mapel", "kategori_mapel", "tingkat", "jurusan"])
    ws3.append(["Matematika",     "UMUM",      "", ""])
    ws3.append(["Penjasorkes",    "OLAHRAGA",  "", ""])
    ws3.append(["Pemrograman Web","PRODUKTIF", "X","TKJ"])

    ws4 = wb.create_sheet("alokasi_kurikulum")
    ws4.append(["nama_kelas", "nama_mapel", "durasi_jp"])
    ws4.append(["X TKJ 1", "Matematika",     4])
    ws4.append(["X TKJ 1", "Penjasorkes",    2])
    ws4.append(["X TKJ 1", "Pemrograman Web",30])

    ws5 = wb.create_sheet("penugasan_guru")
    ws5.append(["nama_guru", "nama_mapel"])
    ws5.append(["Budi Santoso, S.Pd",  "Matematika"])
    ws5.append(["Siti Aminah, S.Pd",   "Penjasorkes"])
    ws5.append(["Eko Prasetyo, S.Kom", "Pemrograman Web"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="template_sitab.xlsx"'},
    )


# ═══════════════════════════════════════════════
#  Google Sheets sync & Excel Upload
# ═══════════════════════════════════════════════

@app.post("/api/upload")
async def upload_excel(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        return pull_excel_data(contents)
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/teachers/upload")
async def upload_teachers_excel(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        return import_teachers_from_excel(contents)
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/subjects/upload")
async def upload_subjects_excel(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        return import_subjects_from_excel(contents)
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/teacher-subjects/upload")
async def upload_teacher_subjects_excel(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        return import_teacher_subjects_from_excel(contents)
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/sync/pull")
def sync_pull():
    try:
        return pull_master_data()
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/sync/push")
def sync_push():
    try:
        return export_timetable_to_sheet()
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/sync/lms")
def sync_to_lms():
    import urllib.request
    import urllib.error
    import ssl
    from urllib.parse import urlparse
    
    conn = get_db_connection()
    try:
        lms_url = get_setting("lms_api_url", "").strip()
        lms_key = get_setting("lms_api_key", "").strip()
        
        # Prioritaskan endpoint aktif dari tabel lms_endpoints
        active_ep = get_active_lms_endpoint()
        if active_ep:
            lms_url = active_ep["endpoint_url"].strip()
            lms_key = active_ep["bearer_token"].strip()
        
        if not lms_url or not lms_key:
            raise HTTPException(400, "Belum ada endpoint LMS yang aktif! Tambahkan dan pilih endpoint di bagian Daftar Endpoint LMS.")
            
        # Normalisasi URL agar selalu mengarah ke endpoint API /api/v1/sync-all
        # Mengatasi kasus jika user menginput domain utama atau URL halaman admin login
        parsed_url = urlparse(lms_url)
        scheme = parsed_url.scheme or "https"
        netloc = parsed_url.netloc
        if not netloc and parsed_url.path:
            parts = parsed_url.path.split('/')
            netloc = parts[0]
        lms_url = f"{scheme}://{netloc}/api/v1/sync-all"
            
        teachers = db_fetchall(conn, "SELECT * FROM teachers ORDER BY nama_guru")
        for t in teachers:
            t["hari_tersedia"]       = json.loads(t["hari_tersedia"] or "[]")
            t["shift_pagi"]          = bool(t["shift_pagi"])
            t["shift_siang"]         = bool(t["shift_siang"])
            t["hari_tersedia_pagi"]  = json.loads(t["hari_tersedia_pagi"])  if t.get("hari_tersedia_pagi")  else None
            t["hari_tersedia_siang"] = json.loads(t["hari_tersedia_siang"]) if t.get("hari_tersedia_siang") else None
            t["min_jp"]              = t.get("min_jp")
            t["max_jp"]              = t.get("max_jp")
            t["allowed_jp_pagi"]     = json.loads(t["allowed_jp_pagi"])     if t.get("allowed_jp_pagi")     else None
            t["allowed_jp_siang"]    = json.loads(t["allowed_jp_siang"])    if t.get("allowed_jp_siang")    else None
            
        classes = db_fetchall(conn, "SELECT nama_kelas, shift_operasional, tingkat, jurusan FROM classes ORDER BY nama_kelas")
        subjects = db_fetchall(conn, "SELECT nama_mapel, kategori_mapel, tingkat, jurusan FROM subjects ORDER BY nama_mapel")
        
        teacher_subjects = db_fetchall(conn, """
            SELECT t.kode_guru, s.nama_mapel
            FROM teacher_subjects ts
            JOIN teachers t ON ts.id_guru = t.id_guru
            JOIN subjects s ON ts.id_mapel = s.id_mapel
            ORDER BY t.nama_guru, s.nama_mapel
        """)
        
        class_subjects = db_fetchall(conn, """
            SELECT c.nama_kelas, s.nama_mapel, cs.durasi_jp, t.kode_guru AS kode_guru_mutlak
            FROM class_subjects cs
            JOIN classes c ON cs.id_kelas = c.id_kelas
            JOIN subjects s ON cs.id_mapel = s.id_mapel
            LEFT JOIN teachers t ON cs.id_guru_mutlak = t.id_guru
            ORDER BY c.nama_kelas, s.nama_mapel
        """)
        
        timetable = db_fetchall(conn, """
            SELECT c.nama_kelas, s.nama_mapel, g.kode_guru, t.hari, t.jam_ke,
                   t.is_fallback, gorig.kode_guru AS original_guru_kode
            FROM timetable t
            JOIN class_subjects cs ON t.id_class_subject = cs.id_class_subject
            JOIN classes c ON cs.id_kelas = c.id_kelas
            JOIN subjects s ON cs.id_mapel = s.id_mapel
            LEFT JOIN teachers g ON t.id_guru = g.id_guru
            LEFT JOIN teachers gorig ON t.original_guru_id = gorig.id_guru
            ORDER BY c.nama_kelas, t.hari, t.jam_ke
        """)
        for slot in timetable:
            slot["is_fallback"] = bool(slot["is_fallback"])
            
        payload = {
            "teachers": teachers,
            "classes": classes,
            "subjects": subjects,
            "teacher_subjects": teacher_subjects,
            "class_subjects": class_subjects,
            "timetable": timetable
        }
        
        import datetime
        def default_serializer(obj):
            if isinstance(obj, (datetime.date, datetime.datetime)):
                return obj.isoformat()
            raise TypeError(f"Type {type(obj)} not serializable")

        try:
            ctx = ssl._create_unverified_context()
            req = urllib.request.Request(
                lms_url,
                data=json.dumps(payload, default=default_serializer).encode('utf-8'),
                headers={
                    'Content-Type': 'application/json',
                    'Authorization': f'Bearer {lms_key}',
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
                },
                method='POST'
            )
            with urllib.request.urlopen(req, context=ctx, timeout=30) as response:
                res_body = response.read().decode('utf-8')
                res_json = json.loads(res_body)
                return {"status": "SUCCESS", "message": res_json.get("message", "Sinkronisasi ke LMS berhasil!")}
        except urllib.error.HTTPError as e:
            err_body = e.read().decode('utf-8')
            logger.error(f"HTTPError syncing to LMS: {err_body}")
            try:
                err_json = json.loads(err_body)
                msg = err_json.get("message", str(e))
            except Exception:
                msg = err_body[:200] or str(e)
            raise HTTPException(status_code=e.code, detail=f"LMS Server Error: {msg}")
        except Exception as e:
            logger.exception("Connection error to LMS")
            raise HTTPException(status_code=500, detail=f"Gagal menghubungi server LMS: {e}")
            
    finally:
        conn.close()


# ═══════════════════════════════════════════════
#  Solver
# ═══════════════════════════════════════════════

@app.post("/api/generate")
def run_generator():
    try:
        return generate_timetable()
    except Exception as e:
        logger.exception("Generator error")
        raise HTTPException(500, f"Gagal generate jadwal: {e}")

@app.post("/api/generate/abort")
def abort_generation():
    from backend.solver import interrupt_active_solver
    success = interrupt_active_solver()
    return {"status": "SUCCESS" if success else "FAILED"}

@app.get("/api/generate/status")
def get_generation_status():
    from backend.solver import active_stage, active_solver
    return {
        "is_running": active_solver is not None or active_stage > 0,
        "stage": active_stage
    }



# ═══════════════════════════════════════════════
#  Coverage Warning (Blueprint para 2 Warning System)
# ═══════════════════════════════════════════════

@app.get("/api/coverage")
def get_coverage_warnings():
    conn = get_db_connection()
    try:
        classes  = db_fetchall(conn, "SELECT * FROM classes")
        teachers = db_fetchall(conn, "SELECT * FROM teachers")

        for t in teachers:
            t["hari_tersedia_pagi"]  = json.loads(t["hari_tersedia_pagi"]  or "[]")
            t["hari_tersedia_siang"] = json.loads(t["hari_tersedia_siang"] or "[]")
            t["shift_pagi"]  = bool(t["shift_pagi"])
            t["shift_siang"] = bool(t["shift_siang"])

        warnings = []
        summary  = []

        for shift in ["PAGI", "SIANG"]:
            n_kelas = sum(1 for c in classes if c["shift_operasional"] == shift)
            if n_kelas == 0:
                continue
            for day in DAYS_ORDER:
                if shift == "PAGI":
                    n_guru = sum(1 for t in teachers if t["shift_pagi"] and day in t["hari_tersedia_pagi"])
                else:
                    n_guru = sum(1 for t in teachers if t["shift_siang"] and day in t["hari_tersedia_siang"])

                entry = {
                    "shift": shift, "hari": day,
                    "kelas_aktif": n_kelas, "guru_tersedia": n_guru,
                    "cukup": n_guru >= n_kelas,
                    "kekurangan": max(0, n_kelas - n_guru),
                }
                summary.append(entry)
                if n_guru < n_kelas:
                    warnings.append(
                        f"Shift {shift}, {day}: {n_kelas} kelas aktif, "
                        f"{n_guru} guru tersedia (kekurangan {n_kelas - n_guru})"
                    )

        return {
            "status":   "OK" if not warnings else "WARNING",
            "warnings": warnings,
            "summary":  summary,
        }
    finally:
        conn.close()


# ═══════════════════════════════════════════════
#  Validasi Pre-Flight
# ═══════════════════════════════════════════════

@app.post("/api/validate")
def validate_data():
    conn = get_db_connection()
    try:
        errors   = []
        warnings = []

        classes  = db_fetchall(conn, "SELECT * FROM classes ORDER BY nama_kelas")
        teachers = db_fetchall(conn, "SELECT * FROM teachers ORDER BY nama_guru")

        for t in teachers:
            t["hari_tersedia_pagi"]  = json.loads(t["hari_tersedia_pagi"]  or "[]")
            t["hari_tersedia_siang"] = json.loads(t["hari_tersedia_siang"] or "[]")
            t["shift_pagi"]  = bool(t["shift_pagi"])
            t["shift_siang"] = bool(t["shift_siang"])
            t["allowed_jp_pagi"]     = json.loads(t["allowed_jp_pagi"])     if t.get("allowed_jp_pagi")     else None
            t["allowed_jp_siang"]    = json.loads(t["allowed_jp_siang"])    if t.get("allowed_jp_siang")    else None

        allocations = db_fetchall(conn, """
            SELECT cs.id_class_subject, cs.id_kelas, cs.id_mapel, cs.durasi_jp,
                   c.nama_kelas, c.shift_operasional, s.nama_mapel, s.kategori_mapel
            FROM   class_subjects cs
            JOIN   classes  c ON cs.id_kelas = c.id_kelas
            JOIN   subjects s ON cs.id_mapel  = s.id_mapel
            ORDER BY c.nama_kelas, s.nama_mapel
        """)

        ts_rows = db_fetchall(conn, "SELECT id_guru, id_mapel FROM teacher_subjects")
        ts_set  = {(r["id_guru"], r["id_mapel"]) for r in ts_rows}

        if not classes:
            errors.append("Belum ada data kelas.")
        if not allocations:
            errors.append("Belum ada alokasi kurikulum.")
        if not ts_rows:
            errors.append("Belum ada penugasan guru (teacher_subjects kosong). Isi Tab 5 dulu.")

        subjects = db_fetchall(conn, "SELECT * FROM subjects")
        subjects_map = {s["id_mapel"]: s for s in subjects}

        if not errors:
            # Panggil preflight terpadu
            pre_errs, pre_warns = _preflight(teachers, classes, allocations, ts_set, subjects_map)
            errors.extend(pre_errs)
            warnings.extend(pre_warns)

            # Tambahkan juga peringatan kekurangan coverage harian biasa (bukan blocker keras)
            for shift in ["PAGI", "SIANG"]:
                n_kelas = sum(1 for c in classes if c["shift_operasional"] == shift)
                if n_kelas == 0:
                    continue
                for day in DAYS_ORDER:
                    if shift == "PAGI":
                        n_guru = sum(1 for t in teachers if t["shift_pagi"] and day in t["hari_tersedia_pagi"])
                    else:
                        n_guru = sum(1 for t in teachers if t["shift_siang"] and day in t["hari_tersedia_siang"])
                    if n_guru < n_kelas:
                        warnings.append(
                            f"Coverage - Shift {shift}, {day}: "
                            f"{n_kelas} kelas aktif, {n_guru} guru tersedia "
                            f"(kekurangan {n_kelas - n_guru})."
                        )

        no_qual = sum(1 for a in allocations if not any(mid == a["id_mapel"] for (_, mid) in ts_set))
        summary_lines = [
            f"Total kelas        : {len(classes)}",
            f"Total alokasi      : {len(allocations)} entri",
            f"Relasi guru-mapel  : {len(ts_rows)}",
            f"Alokasi tanpa guru : {no_qual}",
            f"Error ditemukan    : {len(errors)}",
            f"Peringatan         : {len(warnings)}",
        ]

        return {
            "status":   "OK" if not errors else "ERROR",
            "errors":   errors,
            "warnings": warnings,
            "summary":  summary_lines,
            "stats": {
                "total_classes":     len(classes),
                "total_allocations": len(allocations),
                "ts_count":          len(ts_rows),
                "no_qual_count":     no_qual,
                "error_count":       len(errors),
                "warning_count":     len(warnings),
            },
        }
    finally:
        conn.close()


# ═══════════════════════════════════════════════
#  Health Check Dashboard
# ═══════════════════════════════════════════════

@app.get("/api/health")
def get_health():
    """
    Mengembalikan anomali terstruktur per kategori untuk ditampilkan di dashboard.
    Semua isu yang bisa mencegah generate jadwal dikelompokkan berdasarkan tipe.
    """
    conn = get_db_connection()
    try:
        classes  = db_fetchall(conn, "SELECT * FROM classes ORDER BY nama_kelas")
        teachers = db_fetchall(conn, "SELECT * FROM teachers ORDER BY nama_guru")
        subjects = db_fetchall(conn, "SELECT * FROM subjects ORDER BY nama_mapel")

        for t in teachers:
            t["hari_tersedia_pagi"]  = json.loads(t["hari_tersedia_pagi"]  or "[]")
            t["hari_tersedia_siang"] = json.loads(t["hari_tersedia_siang"] or "[]")
            t["shift_pagi"]  = bool(t["shift_pagi"])
            t["shift_siang"] = bool(t["shift_siang"])
            t["allowed_jp_pagi"]     = json.loads(t["allowed_jp_pagi"])     if t.get("allowed_jp_pagi")     else None
            t["allowed_jp_siang"]    = json.loads(t["allowed_jp_siang"])    if t.get("allowed_jp_siang")    else None

        allocations = db_fetchall(conn, """
            SELECT cs.id_class_subject, cs.id_kelas, cs.id_mapel, cs.durasi_jp,
                   c.nama_kelas, c.shift_operasional, s.nama_mapel, s.kategori_mapel
            FROM   class_subjects cs
            JOIN   classes  c ON cs.id_kelas = c.id_kelas
            JOIN   subjects s ON cs.id_mapel  = s.id_mapel
            ORDER BY c.nama_kelas, s.nama_mapel
        """)

        ts_rows = db_fetchall(conn, "SELECT id_guru, id_mapel FROM teacher_subjects")
        ts_set  = {(r["id_guru"], r["id_mapel"]) for r in ts_rows}
        teachers_map = {t["id_guru"]: t for t in teachers}
        subjects_map = {s["id_mapel"]: s for s in subjects}

        issues = {
            "jp_lebih":            [],  # JP > 40 (ERROR - blocker)
            "jp_kurang":           [],  # JP < 40 (WARNING)
            "mapel_tanpa_guru":    [],  # Alokasi kelas tanpa guru layak (ERROR - blocker)
            "mapel_no_assignment": [],  # Mapel aktif tanpa guru sama sekali (ERROR - blocker)
            "guru_tanpa_mapel":    [],  # Guru belum punya penugasan (WARNING)
            "guru_idle_expert":    [],  # Guru punya keahlian mapel aktif tapi shift tidak match (WARNING)
            "kelas_tanpa_alokasi": [],  # Kelas belum punya alokasi JP (WARNING)
            "olahraga_jp_salah":   [],  # Olahraga harus tepat 2 JP (ERROR - blocker)
            "coverage_kritis":     [],  # Hari dengan kekurangan guru parah (WARNING)
            "feasibility_errors":  [],  # Bottleneck guru / kelayakan (ERROR - blocker)
        }

        # ── 1. JP per kelas ──────────────────────────────────────────
        class_jp: dict = {}
        class_alloc_count: dict = {}
        class_mapel_list: dict = {}
        for a in allocations:
            class_jp[a["id_kelas"]]          = class_jp.get(a["id_kelas"], 0) + a["durasi_jp"]
            class_alloc_count[a["id_kelas"]] = class_alloc_count.get(a["id_kelas"], 0) + 1
            class_mapel_list.setdefault(a["id_kelas"], []).append(a["nama_mapel"])

        for c in classes:
            total   = class_jp.get(c["id_kelas"], 0)
            alloc_n = class_alloc_count.get(c["id_kelas"], 0)
            mapels  = class_mapel_list.get(c["id_kelas"], [])
            if alloc_n == 0:
                issues["kelas_tanpa_alokasi"].append({
                    "kelas": c["nama_kelas"],
                    "shift": c["shift_operasional"],
                    "tingkat": c["tingkat"],
                    "jurusan": c["jurusan"],
                })
            elif total > 40:
                issues["jp_lebih"].append({
                    "kelas":    c["nama_kelas"],
                    "shift":    c["shift_operasional"],
                    "total_jp": total,
                    "selisih":  total - 40,
                    "mapels":   mapels,
                })
            elif total < 40:
                issues["jp_kurang"].append({
                    "kelas":    c["nama_kelas"],
                    "shift":    c["shift_operasional"],
                    "total_jp": total,
                    "selisih":  40 - total,
                    "jumlah_mapel": alloc_n,
                })

        # ── 2. Olahraga JP ──────────────────────────────────────────
        for a in allocations:
            k   = (a.get("nama_mapel") or "").upper()
            kat = (a.get("kategori_mapel") or "").upper()
            is_olah = any(kw in k for kw in ("JASMANI","OLAH RAGA","PENJASORKES","OLAHRAGA","PJOK")) or "OLAHRAGA" in kat
            if is_olah and a["durasi_jp"] != 2:
                issues["olahraga_jp_salah"].append({
                    "kelas":      a["nama_kelas"],
                    "shift":      a["shift_operasional"],
                    "nama_mapel": a["nama_mapel"],
                    "jp_saat_ini": a["durasi_jp"],
                    "hint": "Aturan blok 2 JP berturut-turut wajib dipenuhi solver",
                })

        # ── 3. Alokasi tanpa guru layak ─────────────────────────────
        for a in allocations:
            shift = a["shift_operasional"]
            # Guru yang kualifikasi PERSIS mapel ini
            qualified_ids = [
                gid for (gid, mid) in ts_set
                if mid == a["id_mapel"]
                and teachers_map.get(gid)
                and ((shift == "PAGI"  and teachers_map[gid]["shift_pagi"]) or
                     (shift == "SIANG" and teachers_map[gid]["shift_siang"]))
            ]
            # Guru yang punya keahlian mapel ini tapi shift tidak cocok
            wrong_shift_ids = [
                gid for (gid, mid) in ts_set
                if mid == a["id_mapel"]
                and teachers_map.get(gid)
                and not ((shift == "PAGI"  and teachers_map[gid]["shift_pagi"]) or
                         (shift == "SIANG" and teachers_map[gid]["shift_siang"]))
            ]
            wrong_shift_names = [teachers_map[gid]["nama_guru"] for gid in wrong_shift_ids if gid in teachers_map]

            # Cek kategori mapel untuk saran fallback
            any_assigned = any(mid == a["id_mapel"] for (_, mid) in ts_set)
            if not qualified_ids:
                issues["mapel_tanpa_guru"].append({
                    "kelas":          a["nama_kelas"],
                    "shift":          a["shift_operasional"],
                    "nama_mapel":     a["nama_mapel"],
                    "kategori":       a["kategori_mapel"],
                    "durasi_jp":      a["durasi_jp"],
                    "ada_guru_lain_shift": any_assigned,
                    "guru_shift_salah":    wrong_shift_names,
                })

        # ── 4. Mapel di master tanpa guru sama sekali ───────────────
        assigned_mapel_ids = {mid for (_, mid) in ts_set}
        used_mapel_ids     = {a["id_mapel"] for a in allocations}
        # Hitung berapa kelas pakai mapel ini
        mapel_kelas_count: dict = {}
        for a in allocations:
            mapel_kelas_count[a["id_mapel"]] = mapel_kelas_count.get(a["id_mapel"], 0) + 1

        for s in subjects:
            if s["id_mapel"] in used_mapel_ids and s["id_mapel"] not in assigned_mapel_ids:
                issues["mapel_no_assignment"].append({
                    "id_mapel":    s["id_mapel"],
                    "nama_mapel":  s["nama_mapel"],
                    "kategori":    s["kategori_mapel"],
                    "dipakai_kelas": mapel_kelas_count.get(s["id_mapel"], 0),
                })

        # ── 5. Guru tanpa penugasan mapel apapun ────────────────────
        assigned_guru_ids = {gid for (gid, _) in ts_set}
        for t in teachers:
            if t["id_guru"] not in assigned_guru_ids:
                shift_label = []
                if t["shift_pagi"]:  shift_label.append("PAGI")
                if t["shift_siang"]: shift_label.append("SIANG")
                issues["guru_tanpa_mapel"].append({
                    "nama_guru": t["nama_guru"],
                    "kode_guru": t["kode_guru"],
                    "shift":     "/".join(shift_label) if shift_label else "-",
                    "hint":      "Guru tidak akan pernah dijadwalkan oleh solver",
                })

        # ── 6. Guru idle expert: punya keahlian mapel aktif, tapi shift tidak ada kelas ─────
        classes_pagi  = [c for c in classes if c["shift_operasional"] == "PAGI"]
        classes_siang = [c for c in classes if c["shift_operasional"] == "SIANG"]
        for t in teachers:
            if t["id_guru"] not in assigned_guru_ids:
                continue  # sudah masuk guru_tanpa_mapel
            # Cari keahlian mapel yang aktif dipakai di alokasi
            expert_mapels = [
                subjects_map[mid]["nama_mapel"]
                for (gid, mid) in ts_set
                if gid == t["id_guru"] and mid in used_mapel_ids and mid in subjects_map
            ]
            if not expert_mapels:
                continue
            hints = []
            # Guru shift pagi tapi tidak ada kelas pagi
            if t["shift_pagi"] and not classes_pagi:
                hints.append("Guru shift PAGI tapi tidak ada kelas shift PAGI")
            # Guru shift siang tapi tidak ada kelas siang
            if t["shift_siang"] and not classes_siang:
                hints.append("Guru shift SIANG tapi tidak ada kelas shift SIANG")
            # Guru punya keahlian mapel aktif di shift yang ada kelas, tapi hari tersedia 0
            if t["shift_pagi"] and classes_pagi and not t["hari_tersedia_pagi"]:
                hints.append("Guru shift PAGI tidak punya hari tersedia pagi sama sekali")
            if t["shift_siang"] and classes_siang and not t["hari_tersedia_siang"]:
                hints.append("Guru shift SIANG tidak punya hari tersedia siang sama sekali")
            if hints:
                issues["guru_idle_expert"].append({
                    "nama_guru":    t["nama_guru"],
                    "kode_guru":    t["kode_guru"],
                    "shift":        "/".join([
                        ("PAGI" if t["shift_pagi"] else ""),
                        ("SIANG" if t["shift_siang"] else "")
                    ]).strip("/"),
                    "keahlian_mapel": expert_mapels,
                    "hints":        hints,
                })

        # ── 7. Coverage kritis per shift × hari ─────────────────────
        for shift in ["PAGI", "SIANG"]:
            n_kelas = sum(1 for c in classes if c["shift_operasional"] == shift)
            if n_kelas == 0:
                continue
            for day in DAYS_ORDER:
                if shift == "PAGI":
                    n_guru = sum(1 for t in teachers if t["shift_pagi"] and day in t["hari_tersedia_pagi"])
                else:
                    n_guru = sum(1 for t in teachers if t["shift_siang"] and day in t["hari_tersedia_siang"])

                kekurangan = max(0, n_kelas - n_guru)
                pct_kurang = round(kekurangan / n_kelas * 100) if n_kelas else 0

                if kekurangan > 0:
                    # Siapa saja guru yang tersedia di hari ini?
                    guru_tersedia_names = [
                        t["nama_guru"] for t in teachers
                        if (shift == "PAGI" and t["shift_pagi"] and day in t["hari_tersedia_pagi"]) or
                           (shift == "SIANG" and t["shift_siang"] and day in t["hari_tersedia_siang"])
                    ]
                    severity = "KRITIS" if pct_kurang >= 50 else "KURANG"
                    issues["coverage_kritis"].append({
                        "shift":         shift,
                        "hari":          day,
                        "kelas_aktif":   n_kelas,
                        "guru_tersedia": n_guru,
                        "kekurangan":    kekurangan,
                        "pct_kurang":    pct_kurang,
                        "severity":      severity,
                        "guru_hadir":    guru_tersedia_names,
                    })

        # ── 8. Feasibility Errors (deep analysis) ────────────────────
        issues["feasibility_errors"] = _diagnose_infeasibility(teachers, classes, allocations, ts_set, subjects_map)

        # ── Hitung total error blocker ───────────────────────────────
        total_errors   = (len(issues["jp_lebih"]) + len(issues["mapel_tanpa_guru"]) +
                          len(issues["olahraga_jp_salah"]) + len(issues["mapel_no_assignment"]) +
                          len(issues["feasibility_errors"]))
        total_warnings = (len(issues["jp_kurang"]) + len(issues["guru_tanpa_mapel"]) +
                          len(issues["guru_idle_expert"]) + len(issues["kelas_tanpa_alokasi"]) +
                          len(issues["coverage_kritis"]))

        can_generate = (total_errors == 0 and len(classes) > 0
                        and len(allocations) > 0 and len(ts_rows) > 0)

        return {
            "status":          "OK" if can_generate else ("ERROR" if total_errors > 0 else "WARNING"),
            "can_generate":    can_generate,
            "total_errors":    total_errors,
            "total_warnings":  total_warnings,
            "issues":          issues,
            "summary": {
                "total_kelas":     len(classes),
                "total_guru":      len(teachers),
                "total_mapel":     len(subjects),
                "total_alokasi":   len(allocations),
                "total_penugasan": len(ts_rows),
            },
        }
    finally:
        conn.close()


@app.get("/api/feasibility")
def get_feasibility():
    """
    Menghitung kelayakan ketersediaan guru harian (Daily Coverage)
    dan kelayakan beban mata pelajaran (Subject Capacity).
    """
    conn = get_db_connection()
    try:
        classes = db_fetchall(conn, "SELECT id_kelas, nama_kelas, shift_operasional FROM classes")
        teachers = db_fetchall(conn, "SELECT id_guru, nama_guru, kode_guru, shift_pagi, shift_siang, hari_tersedia_pagi, hari_tersedia_siang, min_jp, max_jp FROM teachers")
        subjects = db_fetchall(conn, "SELECT id_mapel, nama_mapel, kategori_mapel FROM subjects")
        allocations = db_fetchall(conn, """
            SELECT cs.id_class_subject, cs.id_kelas, cs.id_mapel, cs.durasi_jp,
                   c.nama_kelas, c.shift_operasional, s.nama_mapel, s.kategori_mapel
            FROM   class_subjects cs
            JOIN   classes  c ON cs.id_kelas = c.id_kelas
            JOIN   subjects s ON cs.id_mapel  = s.id_mapel
        """)
        ts_rows = db_fetchall(conn, "SELECT id_guru, id_mapel FROM teacher_subjects")
        
        # Parse json fields
        for t in teachers:
            t["hari_tersedia_pagi"]  = json.loads(t["hari_tersedia_pagi"]  or "[]")
            t["hari_tersedia_siang"] = json.loads(t["hari_tersedia_siang"] or "[]")
            t["shift_pagi"]  = bool(t["shift_pagi"])
            t["shift_siang"] = bool(t["shift_siang"])
            t["allowed_jp_pagi"]     = json.loads(t["allowed_jp_pagi"])     if t.get("allowed_jp_pagi")     else None
            t["allowed_jp_siang"]    = json.loads(t["allowed_jp_siang"])    if t.get("allowed_jp_siang")    else None
            
        teachers_map = {t["id_guru"]: t for t in teachers}
        
        # 1. Daily Coverage Map
        daily_coverage = []
        for shift in ["PAGI", "SIANG"]:
            n_kelas = sum(1 for c in classes if c["shift_operasional"] == shift)
            for day in DAYS_ORDER:
                # Guru yang bersedia hadir di shift & hari ini
                guru_hadir = []
                # Guru yang libur (punya shift ini tapi hari ini off)
                guru_libur = []
                
                for t in teachers:
                    if shift == "PAGI" and t["shift_pagi"]:
                        if day in t["hari_tersedia_pagi"]:
                            guru_hadir.append({"id_guru": t["id_guru"], "nama_guru": t["nama_guru"], "kode_guru": t["kode_guru"]})
                        else:
                            guru_libur.append({"id_guru": t["id_guru"], "nama_guru": t["nama_guru"], "kode_guru": t["kode_guru"]})
                    elif shift == "SIANG" and t["shift_siang"]:
                        if day in t["hari_tersedia_siang"]:
                            guru_hadir.append({"id_guru": t["id_guru"], "nama_guru": t["nama_guru"], "kode_guru": t["kode_guru"]})
                        else:
                            guru_libur.append({"id_guru": t["id_guru"], "nama_guru": t["nama_guru"], "kode_guru": t["kode_guru"]})
                
                n_guru = len(guru_hadir)
                kekurangan = max(0, n_kelas - n_guru)
                
                if n_guru < n_kelas:
                    status = "RED"
                elif n_guru <= n_kelas + 2:
                    status = "YELLOW"
                else:
                    status = "GREEN"
                    
                daily_coverage.append({
                    "shift": shift,
                    "hari": day,
                    "butuh": n_kelas,
                    "tersedia": n_guru,
                    "kekurangan": kekurangan,
                    "status": status,
                    "guru_hadir": guru_hadir,
                    "guru_libur": guru_libur
                })

        # 2. Subject Capacity & Shift Breakdown Map
        subject_capacities = []
        recommendations = []

        # Group allocations by subject & shift
        subject_demand_pagi = {}   # mapel_id -> total_jp_pagi
        subject_demand_siang = {}  # mapel_id -> total_jp_siang
        subject_classes_pagi = {}  # mapel_id -> list of class names (pagi)
        subject_classes_siang = {} # mapel_id -> list of class names (siang)
        
        # Allocations with id_guru_mutlak for optimization check
        locked_pagi_teacher_jp = {} # (id_guru, id_kelas) -> total_jp

        for a in allocations:
            mid = a["id_mapel"]
            shift = str(a.get("shift_operasional") or "").strip().upper()
            dur = a["durasi_jp"]
            kname = a["nama_kelas"]

            if shift == "PAGI":
                subject_demand_pagi[mid]  = subject_demand_pagi.get(mid, 0) + dur
                subject_classes_pagi.setdefault(mid, []).append(kname)
            elif shift == "SIANG":
                subject_demand_siang[mid] = subject_demand_siang.get(mid, 0) + dur
                subject_classes_siang.setdefault(mid, []).append(kname)

            if a.get("id_guru_mutlak") and shift == "PAGI":
                key = (a["id_guru_mutlak"], a["id_kelas"], kname)
                locked_pagi_teacher_jp[key] = locked_pagi_teacher_jp.get(key, 0) + dur

        # Group teachers by subject qualifications
        subject_teachers = {} # mapel_id -> list of teachers qualified
        for ts in ts_rows:
            mid = ts["id_mapel"]
            gid = ts["id_guru"]
            if gid in teachers_map:
                subject_teachers.setdefault(mid, []).append(teachers_map[gid])

        for s in subjects:
            mid = s["id_mapel"]
            dem_pagi  = subject_demand_pagi.get(mid, 0)
            dem_siang = subject_demand_siang.get(mid, 0)
            demand    = dem_pagi + dem_siang

            if demand == 0:
                continue # Skip mapel yang tidak dipakai di alokasi kurikulum

            cls_pagi  = list(set(subject_classes_pagi.get(mid, [])))
            cls_siang = list(set(subject_classes_siang.get(mid, [])))
            cls_all   = list(set(cls_pagi + cls_siang))

            qualified_teachers = subject_teachers.get(mid, [])
            teachers_pagi  = [t for t in qualified_teachers if t["shift_pagi"]]
            teachers_siang = [t for t in qualified_teachers if t["shift_siang"]]

            cap_pagi  = sum(t["max_jp"] if t["max_jp"] is not None else 60 for t in teachers_pagi)
            cap_siang = sum(t["max_jp"] if t["max_jp"] is not None else 60 for t in teachers_siang)
            total_capacity = sum(t["max_jp"] if t["max_jp"] is not None else 60 for t in qualified_teachers)

            gurus_info = []
            for t in qualified_teachers:
                gurus_info.append({
                    "id_guru":    t["id_guru"],
                    "nama_guru":  t["nama_guru"],
                    "kode_guru":  t["kode_guru"],
                    "shift_pagi": t["shift_pagi"],
                    "shift_siang":t["shift_siang"],
                    "max_jp":     t["max_jp"] if t["max_jp"] is not None else 60
                })

            if len(qualified_teachers) == 0:
                status = "RED" # Belum ada guru
                recommendations.append({
                    "type": "NO_TEACHER",
                    "severity": "DANGER",
                    "icon": "fa-user-slash",
                    "title": f"Mapel [{s['nama_mapel']}] Belum Punya Guru",
                    "text": f"Mapel [{s['nama_mapel']}] digunakan di {len(cls_all)} kelas (Total {demand} JP), tetapi belum ada guru yang berkualifikasi di Tab 5 (Teacher Subjects)."
                })
            elif dem_pagi > cap_pagi or dem_siang > cap_siang:
                status = "RED" # Kebutuhan melebihi kapasitas per shift
                if dem_pagi > cap_pagi:
                    recommendations.append({
                        "type": "DEFICIENT_SHIFT_PAGI",
                        "severity": "DANGER",
                        "icon": "fa-sun",
                        "title": f"Mapel [{s['nama_mapel']}] Kekurangan Guru (Shift Pagi)",
                        "text": f"Shift PAGI butuh {dem_pagi} JP untuk {len(cls_pagi)} kelas, tetapi kapasitas guru pengampu aktif Pagi hanya {cap_pagi} JP. Tambahkan kualifikasi guru di Tab 5 atau aktifkan shift Pagi guru pengampu."
                    })
                if dem_siang > cap_siang:
                    recommendations.append({
                        "type": "DEFICIENT_SHIFT_SIANG",
                        "severity": "DANGER",
                        "icon": "fa-moon",
                        "title": f"Mapel [{s['nama_mapel']}] Kekurangan Guru (Shift Siang)",
                        "text": f"Shift SIANG butuh {dem_siang} JP untuk {len(cls_siang)} kelas, tetapi kapasitas guru pengampu aktif Siang hanya {cap_siang} JP."
                    })
            elif total_capacity < demand:
                status = "RED"
            elif demand > 0.85 * total_capacity:
                status = "YELLOW"
            else:
                status = "GREEN"

            subject_capacities.append({
                "id_mapel":            mid,
                "nama_mapel":          s["nama_mapel"],
                "kategori_mapel":      s["kategori_mapel"],
                "jp_pagi":             dem_pagi,
                "jp_siang":            dem_siang,
                "total_jp_butuh":      demand,
                "n_kelas_pagi":        len(cls_pagi),
                "n_kelas_siang":       len(cls_siang),
                "n_kelas_total":       len(cls_all),
                "kelas_pemakai_pagi":  cls_pagi,
                "kelas_pemakai_siang": cls_siang,
                "kelas_pemakai":        cls_all,
                "n_guru":              len(qualified_teachers),
                "n_guru_pagi":         len(teachers_pagi),
                "n_guru_siang":        len(teachers_siang),
                "cap_pagi":            cap_pagi,
                "cap_siang":           cap_siang,
                "total_jp_kapasitas":  total_capacity,
                "status":              status,
                "guru_pengampu":       gurus_info,
            })

        # 3. Analisis Rekomendasi Kunci Guru (Guru Mutlak Optimization Check)
        for (gid, cid, kname), locked_dur in locked_pagi_teacher_jp.items():
            g = teachers_map.get(gid)
            if g and locked_dur <= 4:
                recommendations.append({
                    "type": "GURU_MUTLAK_PAGI_LOW",
                    "severity": "WARNING",
                    "icon": "fa-key",
                    "title": f"Saran Formasi: Kunci Guru [{g['nama_guru']}] di Pagi",
                    "text": f"Guru [{g['nama_guru']}] dikunci (Guru Mutlak) di kelas Pagi [{kname}] hanya untuk total {locked_dur} JP. Disarankan untuk tidak dikunci di Pagi agar slot waktu Pagi guru ini bisa dimanfaatkan secara lebih fleksibel."
                })

        return {
            "daily_coverage":     daily_coverage,
            "subject_capacities": subject_capacities,
            "recommendations":    recommendations
        }
    finally:
        conn.close()



# ═══════════════════════════════════════════════
#  Timetable read (JOIN via id_class_subject)
# ═══════════════════════════════════════════════

@app.get("/api/timetable")
def get_timetable():
    conn = get_db_connection()
    try:
        classes = db_fetchall(conn, "SELECT * FROM classes ORDER BY nama_kelas")
        entries = db_fetchall(conn, """
            SELECT t.id_timetable, t.hari, t.jam_ke,
                   t.is_fallback, t.original_guru_id, t.id_class_subject,
                   cs.id_kelas, cs.id_mapel, cs.durasi_jp,
                   c.nama_kelas, c.shift_operasional,
                   s.nama_mapel, s.kategori_mapel,
                   g.id_guru, g.nama_guru, g.kode_guru
            FROM timetable t
            JOIN class_subjects cs ON t.id_class_subject = cs.id_class_subject
            JOIN classes  c ON cs.id_kelas = c.id_kelas
            JOIN subjects s ON cs.id_mapel  = s.id_mapel
            LEFT JOIN teachers g ON t.id_guru = g.id_guru
            ORDER BY c.nama_kelas, t.hari, t.jam_ke
        """)

        for e in entries:
            e["is_fallback"] = 1 if e.get("is_fallback") else 0

        total_possible = sum(
            SHIFT_LIMITS[c["shift_operasional"]][d]
            for c in classes
            for d in SHIFT_LIMITS[c["shift_operasional"]]
        )
        fill_pct = round(len(entries) / total_possible * 100, 1) if total_possible else 0.0

        return {
            "timetable": entries,
            "classes":   classes,
            "stats": {
                "total_slots":          len(entries),
                "total_possible_slots": total_possible,
                "fill_percentage":      fill_pct,
                "classes_count":        len(classes),
            },
        }
    finally:
        conn.close()

@app.delete("/api/timetable")
def delete_timetable():
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        cur.execute("SET FOREIGN_KEY_CHECKS = 0;")
        cur.execute("TRUNCATE TABLE timetable")
        cur.execute("SET FOREIGN_KEY_CHECKS = 1;")
        conn.commit()
        return {"status": "SUCCESS", "message": "Jadwal berhasil direset (dikosongkan)"}
    finally:
        cur.close(); conn.close()


@app.get("/api/timetable/download")
def download_timetable(branch: str = "bekasi"):
    try:
        branch = branch.strip().lower()
        if branch not in ["bekasi", "jakarta"]:
            branch = "bekasi"
        active_branch.set(branch)
        
        buf, filename = generate_excel_timetable()
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        logger.exception("Failed to generate Excel timetable")
        raise HTTPException(500, f"Gagal generate file Excel: {e}")


# ═══════════════════════════════════════════════
#  CRUD - Teachers
# ═══════════════════════════════════════════════

@app.get("/api/teachers", response_model=List[Teacher])
def get_teachers():
    conn = get_db_connection()
    rows = db_fetchall(conn, "SELECT * FROM teachers ORDER BY nama_guru")
    conn.close()
    for d in rows:
        d["hari_tersedia"]       = json.loads(d["hari_tersedia"] or "[]")
        d["shift_pagi"]          = bool(d["shift_pagi"])
        d["shift_siang"]         = bool(d["shift_siang"])
        d["hari_tersedia_pagi"]  = json.loads(d["hari_tersedia_pagi"])  if d.get("hari_tersedia_pagi") is not None else d["hari_tersedia"]
        d["hari_tersedia_siang"] = json.loads(d["hari_tersedia_siang"]) if d.get("hari_tersedia_siang") is not None else d["hari_tersedia"]
        d["min_jp"]              = d.get("min_jp")
        d["max_jp"]              = d.get("max_jp")
        d["allowed_jp_pagi"]     = json.loads(d["allowed_jp_pagi"])     if d.get("allowed_jp_pagi")     else None
        d["allowed_jp_siang"]    = json.loads(d["allowed_jp_siang"])    if d.get("allowed_jp_siang")    else None
    return rows

@app.post("/api/teachers", response_model=Teacher)
def create_teacher(body: TeacherCreate):
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO teachers
                (nama_guru, kode_guru, hari_tersedia, shift_pagi, shift_siang,
                 hari_tersedia_pagi, hari_tersedia_siang, min_jp, max_jp,
                 allowed_jp_pagi, allowed_jp_siang, no_wa)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
            body.nama_guru, body.kode_guru,
            json.dumps(body.hari_tersedia),
            body.shift_pagi, body.shift_siang,
            json.dumps(body.hari_tersedia_pagi)  if body.hari_tersedia_pagi  else None,
            json.dumps(body.hari_tersedia_siang) if body.hari_tersedia_siang else None,
            body.min_jp,
            body.max_jp,
            json.dumps(body.allowed_jp_pagi)     if body.allowed_jp_pagi     else None,
            json.dumps(body.allowed_jp_siang)    if body.allowed_jp_siang    else None,
            body.no_wa,
        ))
        conn.commit()
        new_id = cur.lastrowid
        return {**body.model_dump(), "id_guru": new_id}
    except pymysql.err.IntegrityError:
        conn.rollback()
        raise HTTPException(400, f"Kode guru {body.kode_guru} sudah terdaftar!")
    finally:
        cur.close(); conn.close()

@app.put("/api/teachers/availability")
def update_teachers_availability(items: List[TeacherAvailabilityBatchItem]):
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        for item in items:
            pagi_json  = json.dumps(item.hari_tersedia_pagi)
            siang_json = json.dumps(item.hari_tersedia_siang)
            combined_days = list(dict.fromkeys(item.hari_tersedia_pagi + item.hari_tersedia_siang))
            hari_tersedia_json = json.dumps(combined_days)
            cur.execute("""
                UPDATE teachers SET
                    hari_tersedia = %s,
                    hari_tersedia_pagi = %s,
                    hari_tersedia_siang = %s
                WHERE id_guru = %s
            """, (hari_tersedia_json, pagi_json, siang_json, item.id_guru))
        conn.commit()
        return {"status": "SUCCESS", "updated_count": len(items)}
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"Gagal update ketersediaan hari: {e}")
    finally:
        cur.close(); conn.close()

@app.put("/api/teachers/{id_guru}/availability")
def update_single_teacher_availability(id_guru: int, body: TeacherAvailabilityBatchItem):
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        pagi_json  = json.dumps(body.hari_tersedia_pagi)
        siang_json = json.dumps(body.hari_tersedia_siang)
        combined_days = list(dict.fromkeys(body.hari_tersedia_pagi + body.hari_tersedia_siang))
        hari_tersedia_json = json.dumps(combined_days)
        cur.execute("""
            UPDATE teachers SET
                hari_tersedia = %s,
                hari_tersedia_pagi = %s,
                hari_tersedia_siang = %s
            WHERE id_guru = %s
        """, (hari_tersedia_json, pagi_json, siang_json, id_guru))
        conn.commit()
        return {"status": "SUCCESS", "id_guru": id_guru}
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"Gagal update ketersediaan hari: {e}")
    finally:
        cur.close(); conn.close()

@app.put("/api/teachers/{id_guru}")
def update_teacher(id_guru: int, body: TeacherCreate):
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        cur.execute("""
            UPDATE teachers SET
                nama_guru = %s, kode_guru = %s,
                hari_tersedia = %s, shift_pagi = %s, shift_siang = %s,
                hari_tersedia_pagi = %s, hari_tersedia_siang = %s,
                min_jp = %s, max_jp = %s,
                allowed_jp_pagi = %s, allowed_jp_siang = %s,
                no_wa = %s
            WHERE id_guru = %s
        """, (
            body.nama_guru, body.kode_guru,
            json.dumps(body.hari_tersedia),
            body.shift_pagi, body.shift_siang,
            json.dumps(body.hari_tersedia_pagi)  if body.hari_tersedia_pagi  else None,
            json.dumps(body.hari_tersedia_siang) if body.hari_tersedia_siang else None,
            body.min_jp,
            body.max_jp,
            json.dumps(body.allowed_jp_pagi)     if body.allowed_jp_pagi     else None,
            json.dumps(body.allowed_jp_siang)    if body.allowed_jp_siang    else None,
            body.no_wa,
            id_guru,
        ))
        conn.commit()
        return {**body.model_dump(), "id_guru": id_guru}
    except pymysql.err.IntegrityError:
        conn.rollback()
        raise HTTPException(400, f"Kode guru {body.kode_guru} sudah terdaftar!")
    finally:
        cur.close(); conn.close()

@app.delete("/api/teachers/{id_guru}")
def delete_teacher(id_guru: int):
    conn = get_db_connection()
    cur  = conn.cursor()
    cur.execute("DELETE FROM teachers WHERE id_guru = %s", (id_guru,))
    conn.commit(); cur.close(); conn.close()
    return {"status": "SUCCESS", "message": "Guru berhasil dihapus"}


# ═══════════════════════════════════════════════
#  CRUD - Classes
# ═══════════════════════════════════════════════

def _auto_tingkat_jurusan(nama_kelas: str):
    parts = nama_kelas.split()
    if len(parts) >= 2:
        tk = parts[0].upper()
        if tk in ("X", "XI", "XII"):
            return tk, parts[1].upper()
    return None, None

@app.get("/api/classes", response_model=List[Class])
def get_classes():
    conn = get_db_connection()
    rows = db_fetchall(conn, "SELECT * FROM classes ORDER BY nama_kelas")
    conn.close()
    return rows

@app.post("/api/classes", response_model=Class)
def create_class(body: ClassCreate):
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        nama  = body.nama_kelas.strip()
        shift = body.shift_operasional.upper()
        tk, jr = _auto_tingkat_jurusan(nama)
        tingkat = body.tingkat or tk
        jurusan = body.jurusan or jr
        cur.execute(
            "INSERT INTO classes (nama_kelas, shift_operasional, tingkat, jurusan) VALUES (%s,%s,%s,%s) ",
            (nama, shift, tingkat, jurusan)
        )
        conn.commit()
        new_id = cur.lastrowid
        return {**body.model_dump(), "id_kelas": new_id, "tingkat": tingkat, "jurusan": jurusan}
    except pymysql.err.IntegrityError:
        conn.rollback()
        raise HTTPException(400, "Nama kelas sudah terdaftar!")
    finally:
        cur.close(); conn.close()

@app.patch("/api/classes/{id_kelas}")
def update_class(id_kelas: int, body: ClassCreate):
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        nama  = body.nama_kelas.strip()
        shift = body.shift_operasional.upper()
        tk, jr = _auto_tingkat_jurusan(nama)
        tingkat = body.tingkat or tk
        jurusan = body.jurusan or jr
        cur.execute("""
            UPDATE classes SET nama_kelas=%s, shift_operasional=%s, tingkat=%s, jurusan=%s
            WHERE id_kelas=%s
        """, (nama, shift, tingkat, jurusan, id_kelas))
        conn.commit()
        
        return {**body.model_dump(), "id_kelas": id_kelas, "tingkat": tingkat, "jurusan": jurusan}
    except pymysql.err.IntegrityError:
        conn.rollback()
        raise HTTPException(400, "Nama kelas sudah terdaftar!")
    finally:
        cur.close(); conn.close()

@app.delete("/api/classes/{id_kelas}")
def delete_class(id_kelas: int):
    conn = get_db_connection()
    cur  = conn.cursor()
    cur.execute("DELETE FROM classes WHERE id_kelas = %s", (id_kelas,))
    conn.commit(); cur.close(); conn.close()
    return {"status": "SUCCESS", "message": "Kelas berhasil dihapus"}


# ═══════════════════════════════════════════════
#  CRUD - Subjects
# ═══════════════════════════════════════════════

@app.get("/api/subjects", response_model=List[Subject])
def get_subjects(id_kelas: Optional[int] = None):
    conn = get_db_connection()
    try:
        if id_kelas:
            kelas = db_fetchone(conn, "SELECT tingkat, jurusan FROM classes WHERE id_kelas = %s", (id_kelas,))
            if kelas:
                sql, params = "SELECT * FROM subjects WHERE 1=1", []
                if kelas.get("jurusan"):
                    sql += " AND (jurusan IS NULL OR jurusan = '' OR UPPER(jurusan) = UPPER(%s))"
                    params.append(kelas["jurusan"])
                if kelas.get("tingkat"):
                    sql += " AND (tingkat IS NULL OR tingkat = '' OR UPPER(tingkat) = UPPER(%s))"
                    params.append(kelas["tingkat"])
                sql += " ORDER BY nama_mapel"
                return db_fetchall(conn, sql, tuple(params))
        return db_fetchall(conn, "SELECT * FROM subjects ORDER BY nama_mapel")
    finally:
        conn.close()

@app.post("/api/subjects", response_model=Subject)
def create_subject(body: SubjectCreate):
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO subjects (nama_mapel, kategori_mapel, tingkat, jurusan) VALUES (%s,%s,%s,%s) ",
            (body.nama_mapel.strip(), body.kategori_mapel.upper(),
             body.tingkat.strip().upper() if body.tingkat else None,
             body.jurusan.strip().upper() if body.jurusan else None)
        )
        conn.commit()
        new_id = cur.lastrowid
        return {**body.model_dump(), "id_mapel": new_id}
    finally:
        cur.close(); conn.close()

@app.patch("/api/subjects/{id_mapel}")
def update_subject(id_mapel: int, body: SubjectCreate):
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        cur.execute("""
            UPDATE subjects SET nama_mapel=%s, kategori_mapel=%s, tingkat=%s, jurusan=%s
            WHERE id_mapel=%s
        """, (
            body.nama_mapel.strip(), body.kategori_mapel.upper(),
            body.tingkat.strip().upper() if body.tingkat else None,
            body.jurusan.strip().upper() if body.jurusan else None,
            id_mapel,
        ))
        conn.commit()
        
        return {**body.model_dump(), "id_mapel": id_mapel}
    finally:
        cur.close(); conn.close()

@app.delete("/api/subjects/{id_mapel}")
def delete_subject(id_mapel: int):
    conn = get_db_connection()
    cur  = conn.cursor()
    cur.execute("DELETE FROM subjects WHERE id_mapel = %s", (id_mapel,))
    conn.commit(); cur.close(); conn.close()
    return {"status": "SUCCESS", "message": "Mata Pelajaran berhasil dihapus"}


# ═══════════════════════════════════════════════
#  CRUD - Class Subjects (Alokasi - tanpa id_guru)
# ═══════════════════════════════════════════════

@app.get("/api/allocations", response_model=List[ClassSubject])
def get_allocations():
    conn = get_db_connection()
    rows = db_fetchall(conn, """
        SELECT cs.id_class_subject, cs.id_kelas, cs.id_mapel, cs.durasi_jp, cs.id_guru_mutlak,
               c.nama_kelas, s.nama_mapel, tg.nama_guru AS nama_guru_mutlak
        FROM   class_subjects cs
        JOIN   classes  c ON cs.id_kelas = c.id_kelas
        JOIN   subjects s ON cs.id_mapel  = s.id_mapel
        LEFT JOIN teachers tg ON cs.id_guru_mutlak = tg.id_guru
        ORDER BY c.nama_kelas, s.nama_mapel
    """)
    conn.close()
    return rows

@app.post("/api/allocations", response_model=ClassSubject)
def create_allocation(body: ClassSubjectCreate):
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO class_subjects (id_kelas, id_mapel, durasi_jp, id_guru_mutlak) VALUES (%s,%s,%s,%s) ",
            (body.id_kelas, body.id_mapel, body.durasi_jp, body.id_guru_mutlak)
        )
        conn.commit()
        new_id = cur.lastrowid
        row = db_fetchone(conn, """
            SELECT cs.id_class_subject, cs.id_kelas, cs.id_mapel, cs.durasi_jp, cs.id_guru_mutlak,
                   c.nama_kelas, s.nama_mapel, tg.nama_guru AS nama_guru_mutlak
            FROM   class_subjects cs
            JOIN   classes  c ON cs.id_kelas = c.id_kelas
            JOIN   subjects s ON cs.id_mapel  = s.id_mapel
            LEFT JOIN teachers tg ON cs.id_guru_mutlak = tg.id_guru
            WHERE  cs.id_class_subject = %s
        """, (new_id,))
        return row
    except pymysql.err.IntegrityError:
        conn.rollback()
        raise HTTPException(400, "Mata pelajaran sudah dialokasikan untuk kelas ini!")
    finally:
        cur.close(); conn.close()

@app.delete("/api/allocations/{id_class_subject}")
def delete_allocation(id_class_subject: int):
    conn = get_db_connection()
    cur  = conn.cursor()
    cur.execute("DELETE FROM class_subjects WHERE id_class_subject = %s", (id_class_subject,))
    conn.commit(); cur.close(); conn.close()
    return {"status": "SUCCESS", "message": "Alokasi berhasil dihapus"}

@app.patch("/api/allocations/{id_class_subject}")
def update_allocation(id_class_subject: int, body: AllocationUpdate):
    if body.durasi_jp <= 0 or body.durasi_jp > 40:
        raise HTTPException(400, "Durasi JP harus di antara 1 dan 40 JP")
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        cur.execute("""
            UPDATE class_subjects 
            SET durasi_jp = %s, id_guru_mutlak = %s
            WHERE id_class_subject = %s 
            """, (body.durasi_jp, body.id_guru_mutlak, id_class_subject))
        conn.commit()
        
        return {"status": "SUCCESS", "message": "Alokasi berhasil diperbarui"}
    finally:
        cur.close(); conn.close()


@app.post("/api/allocations/copy")
def copy_allocations(body: AllocationCopy):
    id_asal = body.id_kelas_asal
    id_tujuan = body.id_kelas_tujuan
    if id_asal == id_tujuan:
        raise HTTPException(400, "Kelas asal dan kelas tujuan tidak boleh sama!")
    
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        # Check if source class exists
        cur.execute("SELECT id_kelas FROM classes WHERE id_kelas = %s", (id_asal,))
        if not cur.fetchone():
            raise HTTPException(404, "Kelas asal tidak ditemukan!")
        
        # Check if target class exists
        cur.execute("SELECT id_kelas FROM classes WHERE id_kelas = %s", (id_tujuan,))
        if not cur.fetchone():
            raise HTTPException(404, "Kelas tujuan tidak ditemukan!")
        
        # Get source allocations
        cur.execute("SELECT id_mapel, durasi_jp, id_guru_mutlak FROM class_subjects WHERE id_kelas = %s", (id_asal,))
        source_allocs = cur.fetchall()
        
        if not source_allocs:
            raise HTTPException(400, "Kelas asal tidak memiliki alokasi kurikulum!")
        
        # Clear existing allocations for target class
        cur.execute("DELETE FROM class_subjects WHERE id_kelas = %s", (id_tujuan,))
        
        # Insert duplicated allocations
        for alloc in source_allocs:
            cur.execute(
                "INSERT INTO class_subjects (id_kelas, id_mapel, durasi_jp, id_guru_mutlak) VALUES (%s, %s, %s, %s)",
                (id_tujuan, alloc["id_mapel"], alloc["durasi_jp"], alloc["id_guru_mutlak"])
            )
        conn.commit()
        return {"status": "SUCCESS", "message": f"Berhasil menyalin {len(source_allocs)} alokasi kurikulum."}
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"Gagal menyalin alokasi: {str(e)}")
    finally:
        cur.close(); conn.close()


# ═══════════════════════════════════════════════
#  CRUD - Teacher Subjects (Penugasan Guru)
# ═══════════════════════════════════════════════

@app.get("/api/teacher-subjects", response_model=List[TeacherSubject])
def get_teacher_subjects():
    conn = get_db_connection()
    rows = db_fetchall(conn, """
        SELECT ts.id_teacher_subject, ts.id_guru, ts.id_mapel,
               t.nama_guru, t.kode_guru, s.nama_mapel
        FROM   teacher_subjects ts
        JOIN   teachers t ON ts.id_guru  = t.id_guru
        JOIN   subjects s ON ts.id_mapel = s.id_mapel
        ORDER BY t.nama_guru, s.nama_mapel
    """)
    conn.close()
    return rows

@app.post("/api/teacher-subjects", response_model=TeacherSubject)
def create_teacher_subject(body: TeacherSubjectCreate):
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO teacher_subjects (id_guru, id_mapel) VALUES (%s,%s) ",
            (body.id_guru, body.id_mapel)
        )
        conn.commit()
        new_id = cur.lastrowid
        row = db_fetchone(conn, """
            SELECT ts.id_teacher_subject, ts.id_guru, ts.id_mapel,
                   t.nama_guru, t.kode_guru, s.nama_mapel
            FROM   teacher_subjects ts
            JOIN   teachers t ON ts.id_guru  = t.id_guru
            JOIN   subjects s ON ts.id_mapel = s.id_mapel
            WHERE  ts.id_teacher_subject = %s
        """, (new_id,))
        return row
    except pymysql.err.IntegrityError:
        conn.rollback()
        raise HTTPException(400, "Penugasan guru untuk mata pelajaran ini sudah ada!")
    finally:
        cur.close(); conn.close()

@app.delete("/api/teacher-subjects/{id_teacher_subject}")
def delete_teacher_subject(id_teacher_subject: int):
    conn = get_db_connection()
    cur  = conn.cursor()
    cur.execute("DELETE FROM teacher_subjects WHERE id_teacher_subject = %s", (id_teacher_subject,))
    conn.commit(); cur.close(); conn.close()
    return {"status": "SUCCESS", "message": "Penugasan guru berhasil dihapus"}


# ═══════════════════════════════════════════════
#  Dashboard stats
# ═══════════════════════════════════════════════

@app.get("/api/stats")
def get_stats():
    conn = get_db_connection()
    try:
        n_t  = db_fetchone(conn, "SELECT COUNT(*) AS n FROM teachers")["n"]
        n_c  = db_fetchone(conn, "SELECT COUNT(*) AS n FROM classes")["n"]
        n_s  = db_fetchone(conn, "SELECT COUNT(*) AS n FROM subjects")["n"]
        n_a  = db_fetchone(conn, "SELECT COUNT(*) AS n FROM class_subjects")["n"]
        n_ts = db_fetchone(conn, "SELECT COUNT(*) AS n FROM teacher_subjects")["n"]
        n_tt = db_fetchone(conn, "SELECT COUNT(*) AS n FROM timetable")["n"]
        n_fb = db_fetchone(conn, "SELECT COUNT(*) AS n FROM timetable WHERE is_fallback = TRUE")["n"]
        return {
            "teachers":         n_t,
            "classes":          n_c,
            "subjects":         n_s,
            "allocations":      n_a,
            "teacher_subjects": n_ts,
            "timetable_slots":  n_tt,
            "fallback_count":   n_fb,
        }
    finally:
        conn.close()


# ═══════════════════════════════════════════════
#  Time Slots (Pengaturan Jam Pelajaran & Waktu)
# ═══════════════════════════════════════════════

@app.get("/api/time-slots")
def get_time_slots(shift: Optional[str] = None, hari: Optional[str] = None):
    conn = get_db_connection()
    try:
        query = "SELECT * FROM time_slots WHERE 1=1"
        params = []
        if shift:
            query += " AND shift = %s"
            params.append(shift.upper())
        if hari:
            query += " AND hari = %s"
            params.append(hari.upper())
        query += " ORDER BY FIELD(hari, 'SENIN','SELASA','RABU','KAMIS','JUMAT','SABTU'), urutan ASC, jam_ke ASC"
        slots = db_fetchall(conn, query, tuple(params))
        return slots
    finally:
        conn.close()

@app.post("/api/time-slots/bulk")
def save_time_slots_bulk(payload: TimeSlotBulkSave):
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        hari  = payload.hari.upper()
        shift = payload.shift.upper()
        
        # Hapus slot lama untuk hari & shift ini
        cur.execute("DELETE FROM time_slots WHERE hari = %s AND shift = %s", (hari, shift))
        
        # Insert slot baru
        insert_data = []
        for idx, item in enumerate(payload.slots, start=1):
            insert_data.append((
                hari,
                shift,
                item.jam_ke if item.tipe_slot == 'KBM' else None,
                item.tipe_slot.upper(),
                item.jam_mulai,
                item.jam_selesai,
                item.keterangan or ('Jam Ke-' + str(item.jam_ke) if item.jam_ke else item.tipe_slot),
                item.urutan or idx
            ))
        
        if insert_data:
            cur.executemany(
                "INSERT INTO time_slots (hari, shift, jam_ke, tipe_slot, jam_mulai, jam_selesai, keterangan, urutan) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                insert_data
            )
        
        conn.commit()
        return {"status": "SUCCESS", "message": f"Waktu slot untuk {hari} ({shift}) berhasil disimpan", "count": len(insert_data)}
    except Exception as e:
        conn.rollback()
        logger.error(f"Gagal menyimpan time slots: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

@app.post("/api/time-slots/copy")
def copy_time_slots(payload: TimeSlotCopy):
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        hari_asal = payload.hari_asal.upper()
        shift     = payload.shift.upper()
        
        # Ambil slot sumber
        cur.execute("SELECT * FROM time_slots WHERE hari = %s AND shift = %s ORDER BY urutan ASC", (hari_asal, shift))
        source_slots = cur.fetchall()
        
        if not source_slots:
            raise HTTPException(status_code=400, detail=f"Tidak ada alokasi waktu di hari {hari_asal}")
            
        for h_tujuan in payload.hari_tujuan:
            h_tujuan_clean = h_tujuan.upper()
            if h_tujuan_clean == hari_asal:
                continue
            cur.execute("DELETE FROM time_slots WHERE hari = %s AND shift = %s", (h_tujuan_clean, shift))
            
            new_data = [
                (h_tujuan_clean, shift, slot['jam_ke'], slot['tipe_slot'], slot['jam_mulai'], slot['jam_selesai'], slot['keterangan'], slot['urutan'])
                for slot in source_slots
            ]
            cur.executemany(
                "INSERT INTO time_slots (hari, shift, jam_ke, tipe_slot, jam_mulai, jam_selesai, keterangan, urutan) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                new_data
            )
            
        conn.commit()
        return {"status": "SUCCESS", "message": f"Alokasi waktu berhasil disalin dari {hari_asal} ke {', '.join(payload.hari_tujuan)}"}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Gagal salin time slots: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

@app.post("/api/time-slots/reset")
def reset_time_slots():
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        cur.execute("TRUNCATE TABLE time_slots;")
        from backend.database import _seed_default_time_slots
        _seed_default_time_slots(cur)
        conn.commit()
        return {"status": "SUCCESS", "message": "Alokasi waktu jam pelajaran berhasil di-reset ke default"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

@app.delete("/api/time-slots/{id_slot}")
def delete_time_slot(id_slot: int):
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        cur.execute("DELETE FROM time_slots WHERE id_slot = %s", (id_slot,))
        conn.commit()
        return {"status": "SUCCESS", "message": "Slot jam berhasil dihapus"}
    finally:
        cur.close()
        conn.close()


# ═══════════════════════════════════════════════
#  Static files (frontend)
# ═══════════════════════════════════════════════

app.mount("/", StaticFiles(directory="frontend", html=True), name="frontend")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
