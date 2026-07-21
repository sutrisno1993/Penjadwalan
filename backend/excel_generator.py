import io
import json
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from backend.database import get_db_connection, db_fetchall, get_setting, active_branch

logger = logging.getLogger(__name__)

DAYS_ORDER = ["SENIN", "SELASA", "RABU", "KAMIS", "JUMAT", "SABTU"]
SHIFT_LIMITS = {
    "PAGI":  {"SENIN": 7, "SELASA": 7, "RABU": 7, "KAMIS": 7, "JUMAT": 6, "SABTU": 7},
    "SIANG": {"SENIN": 7, "SELASA": 7, "RABU": 7, "KAMIS": 7, "JUMAT": 6, "SABTU": 6},
}

# ─────────────────────────────────────────────
# Styling Helpers
# ─────────────────────────────────────────────

def style_range(ws, start_row, start_col, end_row, end_col, border=None, fill=None, alignment=None, font=None):
    """Mengaplikasikan style ke seluruh sel dalam range tertentu (termasuk sel yang akan di-merge)."""
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            if border is not None:
                cell.border = border
            if fill is not None:
                cell.fill = fill
            if alignment is not None:
                cell.alignment = alignment
            if font is not None:
                cell.font = font

def get_wall_clock_time(day, shift, jp, conn=None):
    """Mendapatkan jam pelajaran sesuai dengan data time_slots atau fallback standard."""
    if conn:
        try:
            cur = conn.cursor()
            cur.execute(
                "SELECT jam_mulai, jam_selesai FROM time_slots "
                "WHERE hari = %s AND shift = %s AND jam_ke = %s AND tipe_slot = 'KBM' LIMIT 1",
                (day.upper(), shift.upper(), jp)
            )
            row = cur.fetchone()
            cur.close()
            if row:
                return f"{row['jam_mulai']} - {row['jam_selesai']}"
        except Exception:
            pass

    # Fallback jika belum terisi di DB
    if shift == "PAGI":
        if day == "JUMAT":
            slots = ["07:00 - 07:40", "07:40 - 08:20", "08:20 - 09:00", "09:00 - 09:40", "10:10 - 10:50", "10:50 - 11:30"]
            return slots[jp - 1] if 1 <= jp <= len(slots) else ""
        elif day == "SENIN":
            slots = ["07:30 - 08:10", "08:10 - 08:50", "08:50 - 09:30", "10:00 - 10:35", "10:35 - 11:10", "11:10 - 11:45", "11:45 - 12:20"]
            return slots[jp - 1] if 1 <= jp <= len(slots) else ""
        else:
            slots = ["07:00 - 07:45", "07:45 - 08:30", "08:30 - 09:15", "09:15 - 10:00", "10:30 - 11:15", "11:15 - 12:00", "12:00 - 12:45"]
            return slots[jp - 1] if 1 <= jp <= len(slots) else ""
    else:
        slots = ["13:00 - 13:40", "13:40 - 14:20", "14:20 - 15:00", "15:30 - 16:10", "16:10 - 16:50", "16:50 - 17:30"]
        return slots[jp - 1] if 1 <= jp <= len(slots) else ""

def get_break_time(day, shift, conn=None):
    """Mendapatkan jam istirahat sesuai dengan hari dan shift dari database/fallback."""
    if conn:
        try:
            cur = conn.cursor()
            cur.execute(
                "SELECT jam_mulai, jam_selesai FROM time_slots "
                "WHERE hari = %s AND shift = %s AND tipe_slot = 'ISTIRAHAT' LIMIT 1",
                (day.upper(), shift.upper())
            )
            row = cur.fetchone()
            cur.close()
            if row:
                return f"{row['jam_mulai']} - {row['jam_selesai']}"
        except Exception:
            pass

    if shift == "PAGI":
        if day == "SENIN":
            return "09:30 - 10:00"
        elif day == "JUMAT":
            return "09:40 - 10:10"
        else:
            return "10:00 - 10:30"
    else:
        if day == "JUMAT":
            return "15:40 - 16:00"
        else:
            return "15:45 - 16:15"

# ─────────────────────────────────────────────
# Core Generator
# ─────────────────────────────────────────────

def generate_excel_timetable() -> tuple[io.BytesIO, str]:
    """
    Menghasilkan workbook Excel berisi:
      1. Jadwal PAGI (jika ada kelas pagi)
      2. Jadwal SIANG (jika ada kelas siang)
      3. Ringkasan Guru
      4. Alokasi Mengajar (Matriks Guru x Kelas)
    Kembali berupa (bytes_buffer, filename).
    """
    conn = get_db_connection()
    try:
        classes = db_fetchall(conn, "SELECT * FROM classes ORDER BY nama_kelas")
        teachers = db_fetchall(conn, "SELECT * FROM teachers ORDER BY kode_guru")
        subjects = db_fetchall(conn, "SELECT * FROM subjects ORDER BY nama_mapel")
        
        allocations = db_fetchall(conn, """
            SELECT cs.id_class_subject, cs.id_kelas, cs.id_mapel, cs.durasi_jp, cs.id_guru_mutlak,
                   c.nama_kelas, c.shift_operasional,
                   s.nama_mapel, s.kategori_mapel
            FROM   class_subjects cs
            JOIN   classes  c ON cs.id_kelas = c.id_kelas
            JOIN   subjects s ON cs.id_mapel  = s.id_mapel
        """)
        
        tt_rows = db_fetchall(conn, """
            SELECT t.hari, t.jam_ke, t.id_guru, t.is_fallback,
                   cs.id_kelas, cs.id_mapel, s.nama_mapel,
                   g.kode_guru, g.nama_guru
            FROM timetable t
            JOIN class_subjects cs ON t.id_class_subject = cs.id_class_subject
            JOIN subjects s ON cs.id_mapel = s.id_mapel
            LEFT JOIN teachers g ON t.id_guru = g.id_guru
        """)
    finally:
        conn.close()

    from backend.database import get_branch_name
    branch_name = get_branch_name().upper()
    school_name = f"SMK 11 MARET {branch_name}"
    school_year = f"{datetime.now().getFullYear() if hasattr(datetime.now(), 'getFullYear') else datetime.now().year} - {(datetime.now().getFullYear() if hasattr(datetime.now(), 'getFullYear') else datetime.now().year) + 1}"

    wb = openpyxl.Workbook()
    # Hapus sheet default pertama
    if wb.active:
        wb.remove(wb.active)

    # Styles
    font_family = "Arial"
    
    font_title = Font(name=font_family, size=14, bold=True)
    font_subtitle = Font(name=font_family, size=11, bold=True)
    font_badge = Font(name=font_family, size=20, bold=True)
    
    font_header_label = Font(name=font_family, size=10, bold=True)
    font_header_class = Font(name=font_family, size=10, bold=True)
    
    font_data_bold = Font(name=font_family, size=9, bold=True)
    font_data_normal = Font(name=font_family, size=9, bold=False)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_break = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
    
    border_side_thin = Side(style='thin', color='B0B0B0')
    border_side_medium = Side(style='medium', color='000000')
    
    border_cell = Border(left=border_side_thin, right=border_side_thin, top=border_side_thin, bottom=border_side_thin)
    border_cell_day_bottom = Border(left=border_side_thin, right=border_side_thin, top=border_side_thin, bottom=border_side_medium)

    # Index timetable data: (id_kelas, hari, jam_ke) -> entry
    tt_map = {(r["id_kelas"], r["hari"], r["jam_ke"]): r for r in tt_rows}

    # ═══════════════════════════════════════════════
    #  Sheet 1 & 2: Jadwal PAGI & SIANG
    # ═══════════════════════════════════════════════
    for shift in ["PAGI", "SIANG"]:
        shift_classes = [c for c in classes if c["shift_operasional"] == shift]
        if not shift_classes:
            continue

        # Sort classes alphabetically
        shift_classes = sorted(shift_classes, key=lambda x: x["nama_kelas"])

        ws = wb.create_sheet(title=f"Jadwal {shift}")
        ws.views.sheetView[0].showGridLines = True

        total_cols = 4 + len(shift_classes) * 2
        badge_col = total_cols - 1

        # ── 1. Title Block ──
        ws.append(["JADWAL MATA PELAJARAN"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=badge_col - 1)
        ws.cell(row=1, column=1).font = font_title
        ws.cell(row=1, column=1).alignment = align_center
        ws.row_dimensions[1].height = 25

        ws.append([school_name])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=badge_col - 1)
        ws.cell(row=2, column=1).font = font_subtitle
        ws.cell(row=2, column=1).alignment = align_center
        ws.row_dimensions[2].height = 18

        ws.append([f"TAHUN PELAJARAN {school_year}"])
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=badge_col - 1)
        ws.cell(row=3, column=1).font = font_subtitle
        ws.cell(row=3, column=1).alignment = align_center
        ws.row_dimensions[3].height = 18

        # Large shift badge in the top right corner
        ws.cell(row=1, column=badge_col, value=shift)
        ws.merge_cells(start_row=1, start_column=badge_col, end_row=3, end_column=total_cols)
        ws.cell(row=1, column=badge_col).font = font_badge
        ws.cell(row=1, column=badge_col).alignment = align_center

        # Empty row spacer
        ws.append([])
        ws.row_dimensions[4].height = 10

        # ── 2. Column Headers ──
        # Row 5: Header level 1 (Kelas label)
        hdr_row5 = ["No", "Hari", "Jam Ke-", "Waktu"]
        for cls in shift_classes:
            hdr_row5.extend([cls["nama_kelas"], ""])
        ws.append(hdr_row5)
        ws.row_dimensions[5].height = 24
        
        # Merge 'Kelas' headers
        for i, cls in enumerate(shift_classes):
            c_start = 5 + i * 2
            ws.cell(row=5, column=c_start).alignment = align_center
            ws.cell(row=5, column=c_start).font = font_header_class
            style_range(ws, start_row=5, start_col=c_start, end_row=5, end_col=c_start + 1, border=border_cell, fill=fill_header)
            ws.merge_cells(start_row=5, start_column=c_start, end_row=5, end_column=c_start + 1)

        # Merge No, Hari, Jam Ke-, Waktu vertically across Row 5 & Row 6?
        # Wait, since there is NO sub-header row in the screenshot (no "Mata Pelajaran" / "Guru"), we do NOT need Row 6 header!
        # In the user's screenshot, Row 5 contains:
        # No | Hari | Jam Ke- | Waktu | [Class 1 (merged 2 cols)] | [Class 2 (merged 2 cols)] ...
        # This is exactly what we just appended in Row 5!
        # So we don't need Row 6 header at all, we can start data rows immediately at Row 6!
        # Let's style Row 5 columns A-D
        for c in range(1, 5):
            ws.cell(row=5, column=c).alignment = align_center
            ws.cell(row=5, column=c).font = font_header_label
            ws.cell(row=5, column=c).border = border_cell
            ws.cell(row=5, column=c).fill = fill_header

        # ── 3. Data Rows ──
        row_idx = 6
        day_number = 1

        for day in DAYS_ORDER:
            max_jp = SHIFT_LIMITS[shift].get(day, 0)
            if max_jp == 0:
                continue

            day_start_row = row_idx

            for jp in range(1, max_jp + 1):
                # Apply Upacara flag for Monday PAGI JP 1
                is_upacara = (day == "SENIN" and jp == 1 and shift == "PAGI")

                if is_upacara:
                    # Monday JP 1 (Upacara Bendera) row
                    # Jam Ke- is blank, Waktu is 06:30 - 07:30
                    row_data = ["", "", "", "06.30 - 07.30"]
                    for _ in shift_classes:
                        row_data.extend(["UPACARA BENDERA", ""])
                    ws.append(row_data)
                    ws.row_dimensions[row_idx].height = 24
                    
                    # Style the merged UPACARA BENDERA block for classes
                    style_range(ws, start_row=row_idx, start_col=5, end_row=row_idx, end_col=total_cols, 
                                border=border_cell, fill=fill_break, alignment=align_center, font=font_data_bold)
                    ws.merge_cells(start_row=row_idx, start_column=5, end_row=row_idx, end_column=total_cols)
                    
                    # Style cells A-D
                    for c in range(1, 5):
                        ws.cell(row=row_idx, column=c).border = border_cell
                        ws.cell(row=row_idx, column=c).alignment = align_center
                        ws.cell(row=row_idx, column=c).font = font_data_normal
                    row_idx += 1
                else:
                    # Normal Lesson Row
                    clock = get_wall_clock_time(day, shift, jp)
                    row_data = ["", "", jp, clock]
                    for cls in shift_classes:
                        entry = tt_map.get((cls["id_kelas"], day, jp))
                        if entry:
                            mapel = entry["nama_mapel"]
                            if entry.get("is_fallback"):
                                mapel += " *"
                            kode = entry["kode_guru"] if entry["kode_guru"] is not None else ""
                            row_data.extend([mapel, str(kode)])
                        else:
                            row_data.extend(["", ""])
                    
                    ws.append(row_data)
                    ws.row_dimensions[row_idx].height = 32 # Row height 32 for wrap text spacing!
                    
                    # Apply styles & borders
                    for c in range(1, 5):
                        ws.cell(row=row_idx, column=c).border = border_cell
                        ws.cell(row=row_idx, column=c).alignment = align_center
                        ws.cell(row=row_idx, column=c).font = font_data_normal
                        
                    for i, cls in enumerate(shift_classes):
                        c_mapel = 5 + i * 2
                        c_guru = c_mapel + 1
                        
                        # Subject Cell
                        cell_m = ws.cell(row=row_idx, column=c_mapel)
                        cell_m.border = border_cell
                        cell_m.alignment = align_center_wrap # Center aligned + wrapped text!
                        cell_m.font = font_data_normal
                        
                        # Teacher Code Cell
                        cell_g = ws.cell(row=row_idx, column=c_guru)
                        cell_g.border = border_cell
                        cell_g.alignment = align_center
                        cell_g.font = font_data_normal
                    
                    row_idx += 1

                # ── Break Row (ISTIRAHAT) ──
                # Break is inserted after period 4
                if jp == 4:
                    break_time = get_break_time(day, shift)
                    row_data = ["", "", "", break_time]
                    for _ in shift_classes:
                        row_data.extend(["I S T I R A H A T", ""])
                    ws.append(row_data)
                    ws.row_dimensions[row_idx].height = 20
                    
                    # Merge ISTIRAHAT across class columns
                    style_range(ws, start_row=row_idx, start_col=5, end_row=row_idx, end_col=total_cols, 
                                border=border_cell, fill=fill_break, alignment=align_center, font=font_data_bold)
                    ws.merge_cells(start_row=row_idx, start_column=5, end_row=row_idx, end_column=total_cols)
                    
                    for c in range(1, 5):
                        ws.cell(row=row_idx, column=c).border = border_cell
                        ws.cell(row=row_idx, column=c).alignment = align_center
                        ws.cell(row=row_idx, column=c).font = font_data_normal
                    
                    row_idx += 1

            # Put values in cells (top-left of merged range)
            ws.cell(row=day_start_row, column=1, value=day_number).font = font_header_label
            ws.cell(row=day_start_row, column=1).alignment = align_center
            ws.cell(row=day_start_row, column=2, value=day).font = font_header_label
            ws.cell(row=day_start_row, column=2).alignment = align_center

            # ── Merge No & Hari for this day ──
            day_end_row = row_idx - 1
            if day_end_row > day_start_row:
                ws.merge_cells(start_row=day_start_row, start_column=1, end_row=day_end_row, end_column=1)
                ws.merge_cells(start_row=day_start_row, start_column=2, end_row=day_end_row, end_column=2)
            
            # Apply thick line bottom border to the last row of the day (day separator)
            for c in range(1, total_cols + 1):
                ws.cell(row=day_end_row, column=c).border = Border(
                    left=border_side_thin,
                    right=border_side_thin,
                    top=ws.cell(row=day_end_row, column=c).border.top or border_side_thin,
                    bottom=border_side_medium # thick bottom separator
                )
            
            day_number += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 5   # No
        ws.column_dimensions['B'].width = 10  # Hari
        ws.column_dimensions['C'].width = 7   # Jam Ke-
        ws.column_dimensions['D'].width = 15  # Waktu
        
        for i, cls in enumerate(shift_classes):
            col_letter_m = get_column_letter(5 + i * 2)
            col_letter_g = get_column_letter(6 + i * 2)
            ws.column_dimensions[col_letter_m].width = 18  # Mata Pelajaran (wrap text)
            ws.column_dimensions[col_letter_g].width = 6   # Guru Code

    # ═══════════════════════════════════════════════
    #  Sheet 3: Ringkasan Guru
    # ═══════════════════════════════════════════════
    ws_g = wb.create_sheet(title="Ringkasan Guru")
    ws_g.views.sheetView[0].showGridLines = True

    # Title
    ws_g.append(["RINGKASAN ALOKASI GURU"])
    ws_g.merge_cells("A1:I1")
    ws_g.cell(row=1, column=1).font = font_title
    ws_g.cell(row=1, column=1).alignment = align_center
    ws_g.row_dimensions[1].height = 25

    ws_g.append([school_name])
    ws_g.merge_cells("A2:I2")
    ws_g.cell(row=2, column=1).font = font_subtitle
    ws_g.cell(row=2, column=1).alignment = align_center
    ws_g.row_dimensions[2].height = 18

    ws_g.append([f"TAHUN PELAJARAN {school_year}"])
    ws_g.merge_cells("A3:I3")
    ws_g.cell(row=3, column=1).font = font_subtitle
    ws_g.cell(row=3, column=1).alignment = align_center
    ws_g.row_dimensions[3].height = 18

    ws_g.append([]) # spacer
    ws_g.row_dimensions[4].height = 10

    # Header Row
    headers_g = ["No", "Nama Guru", "Kode", "Shift", "Alokasi Mengajar", "Total JP", "Min JP", "Max JP", "Status"]
    ws_g.append(headers_g)
    ws_g.row_dimensions[5].height = 22
    for c in range(1, 10):
        cell = ws_g.cell(row=5, column=c)
        cell.font = font_header_label
        cell.fill = fill_header
        cell.border = border_cell
        cell.alignment = align_center

    # Fills for status badge
    fill_ok = PatternFill(start_color="DEF7EC", end_color="DEF7EC", fill_type="solid") # light green
    font_ok = Font(name=font_family, size=9, bold=True, color="03543F")
    
    fill_warn = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # light yellow
    font_warn = Font(name=font_family, size=9, bold=True, color="92400E")
    
    fill_err = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid") # light red
    font_err = Font(name=font_family, size=9, bold=True, color="9B1C1C")

    # Fill teacher data rows
    classes_dict = {c["id_kelas"]: c for c in classes}
    subjects_dict = {s["id_mapel"]: s for s in subjects}

    for idx, teacher in enumerate(teachers, start=1):
        tid = teacher["id_guru"]
        t_entries = [e for e in tt_rows if e["id_guru"] == tid]
        
        # Group allocations by class + subject
        combo_map = {}
        for e in t_entries:
            key = (e["id_kelas"], e["id_mapel"])
            combo_map[key] = combo_map.get(key, 0) + 1
            
        alloc_list = []
        for (kid, mid), cnt in combo_map.items():
            k_nama = classes_dict.get(kid, {}).get("nama_kelas", str(kid))
            m_nama = subjects_dict.get(mid, {}).get("nama_mapel", str(mid))
            alloc_list.append(f"{k_nama} - {m_nama} ({cnt} JP)")
            
        alloc_str = " | ".join(alloc_list) if alloc_list else "-"
        total_jp = len(t_entries)
        min_jp = teacher["min_jp"] if teacher.get("min_jp") is not None else 2
        max_jp = teacher["max_jp"] if teacher.get("max_jp") is not None else 60
        
        shift_parts = []
        if teacher.get("shift_pagi"): shift_parts.append("PAGI")
        if teacher.get("shift_siang"): shift_parts.append("SIANG")
        shift_info = ", ".join(shift_parts) if shift_parts else "-"
        
        if total_jp == 0:
            status = "TIDAK MENDAPAT JAM"
            cell_fill, cell_font = fill_err, font_err
        elif total_jp < min_jp:
            status = "KURANG BEBAN"
            cell_fill, cell_font = fill_warn, font_warn
        elif total_jp > max_jp:
            status = "KELEBIHAN BEBAN"
            cell_fill, cell_font = fill_err, font_err
        else:
            status = "OK"
            cell_fill, cell_font = fill_ok, font_ok
            
        row_data = [idx, teacher["nama_guru"], str(teacher["kode_guru"]), shift_info, alloc_str, total_jp, min_jp, max_jp, status]
        ws_g.append(row_data)
        r_idx = idx + 5
        ws_g.row_dimensions[r_idx].height = 20
        
        # Apply styles
        for c in range(1, 10):
            cell = ws_g.cell(row=r_idx, column=c)
            cell.border = border_cell
            cell.font = font_data_normal
            if c in (1, 3, 4, 6, 7, 8):
                cell.alignment = align_center
            elif c == 2:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c == 5:
                cell.alignment = align_left_wrap
            elif c == 9:
                cell.alignment = align_center
                cell.fill = cell_fill
                cell.font = cell_font

    # Column Widths
    ws_g.column_dimensions['A'].width = 5   # No
    ws_g.column_dimensions['B'].width = 28  # Nama Guru
    ws_g.column_dimensions['C'].width = 8   # Kode
    ws_g.column_dimensions['D'].width = 12  # Shift
    ws_g.column_dimensions['E'].width = 60  # Alokasi Mengajar
    ws_g.column_dimensions['F'].width = 10  # Total JP
    ws_g.column_dimensions['G'].width = 8   # Min JP
    ws_g.column_dimensions['H'].width = 8   # Max JP
    ws_g.column_dimensions['I'].width = 20  # Status

    # ═══════════════════════════════════════════════
    #  Sheet 4: Alokasi Mengajar (Matriks)
    # ═══════════════════════════════════════════════
    ws_m = wb.create_sheet(title="Alokasi Mengajar")
    ws_m.views.sheetView[0].showGridLines = True

    # Helper for parsing class levels and majors
    def parse_kelas(nama):
        parts = nama.strip().split()
        tingkat = ""
        jurusan = ""
        rombel = "1"
        if parts and parts[0] in ("X", "XI", "XII"):
            tingkat = parts[0]
            if len(parts) >= 3:
                jurusan = " ".join(parts[1:-1])
                rombel = parts[-1]
            elif len(parts) == 2:
                jurusan = parts[1]
        else:
            jurusan = nama
        return {"tingkat": tingkat, "jurusan": jurusan, "rombel": rombel}

    # Sort classes alphabetically by Level -> Major -> Rombel
    tingkat_order = {"X": 1, "XI": 2, "XII": 3}
    sorted_classes = sorted(classes, key=lambda c: (
        tingkat_order.get(parse_kelas(c["nama_kelas"])["tingkat"], 9),
        parse_kelas(c["nama_kelas"])["jurusan"],
        parse_kelas(c["nama_kelas"])["rombel"]
    ))

    n_classes = len(sorted_classes)
    col_p = 6 + n_classes      # Col index for P (1-based, index 5 is Mapel, then n_classes cols)
    col_s = col_p + 1
    col_jm = col_p + 2
    total_cols_m = col_jm

    # Title
    ws_m.append([f"ALOKASI MENGAJAR {school_name} TAHUN {school_year}"])
    ws_m.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols_m)
    ws_m.cell(row=1, column=1).font = font_title
    ws_m.cell(row=1, column=1).alignment = align_center
    ws_m.row_dimensions[1].height = 25

    ws_m.append([]) # Spacer row
    ws_m.row_dimensions[2].height = 6

    # Build Header level 1 (Tingkat: Kelas X, Kelas XI, Kelas XII)
    hdr1 = ["No", "Nama Guru", "Kode Guru", "Jabatan", "Mapel"]
    tingkat_groups = {}
    for idx, cls in enumerate(sorted_classes):
        p = parse_kelas(cls["nama_kelas"])
        key = p["tingkat"] or "Lainnya"
        if key not in tingkat_groups:
            tingkat_groups[key] = {"start": idx, "count": 0}
        tingkat_groups[key]["count"] += 1
        hdr1.append("")
    hdr1.extend(["P", "S", "JM"])
    ws_m.append(hdr1)
    
    # Merge label tingkat in header Row 3
    for tgt, info in tingkat_groups.items():
        col_start = 6 + info["start"]
        col_end = col_start + info["count"] - 1
        ws_m.cell(row=3, column=col_start, value=f"Kelas {tgt}")
        style_range(ws_m, start_row=3, start_col=col_start, end_row=3, end_col=col_end, border=border_cell, fill=fill_header, font=font_header_label, alignment=align_center)
        if col_start < col_end:
            ws_m.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_end)

    # Build Header level 2 (Jurusan)
    hdr2 = ["", "", "", "", ""]
    jurusan_groups = {}
    for idx, cls in enumerate(sorted_classes):
        p = parse_kelas(cls["nama_kelas"])
        key = (p["tingkat"], p["jurusan"])
        if key not in jurusan_groups:
            jurusan_groups[key] = {"start": idx, "count": 0, "jurusan": p["jurusan"]}
        jurusan_groups[key]["count"] += 1
        hdr2.append("")
    hdr2.extend(["", "", ""])
    ws_m.append(hdr2)
    
    # Merge label jurusan in header Row 4
    for info in jurusan_groups.values():
        col_start = 6 + info["start"]
        col_end = col_start + info["count"] - 1
        ws_m.cell(row=4, column=col_start, value=info["jurusan"])
        style_range(ws_m, start_row=4, start_col=col_start, end_row=4, end_col=col_end, border=border_cell, fill=fill_header, font=font_header_label, alignment=align_center)
        if col_start < col_end:
            ws_m.merge_cells(start_row=4, start_column=col_start, end_row=4, end_column=col_end)

    # Build Header level 3 (Rombel/Nomor kelas)
    hdr3 = ["No", "Nama Guru", "Kode Guru", "Jabatan", "Mapel"]
    for cls in sorted_classes:
        hdr3.append(parse_kelas(cls["nama_kelas"])["rombel"])
    hdr3.extend(["P", "S", "JM"])
    ws_m.append(hdr3)
    
    # Style header Row 5 class numbers
    for idx in range(len(sorted_classes)):
        c_idx = 6 + idx
        ws_m.cell(row=5, column=c_idx).font = font_header_label
        ws_m.cell(row=5, column=c_idx).fill = fill_header
        ws_m.cell(row=5, column=c_idx).border = border_cell
        ws_m.cell(row=5, column=c_idx).alignment = align_center

    # Merge vertical headers (No, Nama Guru, Kode Guru, Jabatan, Mapel, P, S, JM) across Rows 3 to 5
    vertical_headers = [
        (1, "No"), (2, "Nama Guru"), (3, "Kode Guru"), (4, "Jabatan"), (5, "Mapel"),
        (col_p, "P"), (col_s, "S"), (col_jm, "JM")
    ]
    for c_idx, label in vertical_headers:
        ws_m.cell(row=3, column=c_idx, value=label)
        style_range(ws_m, start_row=3, start_col=c_idx, end_row=5, end_col=c_idx, border=border_cell, fill=fill_header, font=font_header_label, alignment=align_center)
        ws_m.merge_cells(start_row=3, start_column=c_idx, end_row=5, end_column=c_idx)

    ws_m.row_dimensions[3].height = 20
    ws_m.row_dimensions[4].height = 18
    ws_m.row_dimensions[5].height = 18

    # Calculate timetable slots and allocations
    # slot_weight[id_guru][id_kelas][id_mapel] = weight (primary vs fallback)
    slot_weight = {}
    for e in tt_rows:
        gid = e["id_guru"]
        kid = e["id_kelas"]
        mid = e["id_mapel"]
        weight = 1 if e["is_fallback"] else 100
        slot_weight.setdefault(gid, {}).setdefault(kid, {})
        slot_weight[gid][kid][mid] = slot_weight[gid][kid].get(mid, 0) + weight

    # Find dominant teacher per (kelas, mapel)
    dominant_teacher = {} # (id_kelas, id_mapel) -> id_guru
    for gid, class_map in slot_weight.items():
        for kid, mapel_map in class_map.items():
            for mid, w in mapel_map.items():
                curr_gid = dominant_teacher.get((kid, mid))
                curr_w = slot_weight.get(curr_gid, {}).get(kid, {}).get(mid, -1) if curr_gid is not None else -1
                if w > curr_w:
                    dominant_teacher[(kid, mid)] = gid

    # Allocations lookup: (id_kelas, id_mapel) -> durasi_jp
    alloc_lookup = {(a["id_kelas"], a["id_mapel"]): a["durasi_jp"] for a in allocations}
    
    # Class total allocations (footer)
    class_total_alloc = {}
    for a in allocations:
        class_total_alloc[a["id_kelas"]] = class_total_alloc.get(a["id_kelas"], 0) + a["durasi_jp"]

    row_idx = 6
    guru_no = 1

    for teacher in teachers:
        tid = teacher["id_guru"]
        
        # Collect subjects for this teacher where they are dominant
        teacher_rows = []
        for i, cls in enumerate(sorted_classes):
            kid = cls["id_kelas"]
            class_allocs = [a for a in allocations if a["id_kelas"] == kid]
            for a in class_allocs:
                mid = a["id_mapel"]
                if dominant_teacher.get((kid, mid)) == tid and a["durasi_jp"] > 0:
                    # Look if this subject row was already added
                    found = False
                    for r in teacher_rows:
                        if r["id_mapel"] == mid:
                            r["jp_per_kelas"][i] = a["durasi_jp"]
                            found = True
                            break
                    if not found:
                        jp_list = [""] * n_classes
                        jp_list[i] = a["durasi_jp"]
                        teacher_rows.append({
                            "id_mapel": mid,
                            "nama_mapel": subjects_dict.get(mid, {}).get("nama_mapel", f"Mapel#{mid}"),
                            "jp_per_kelas": jp_list
                        })

        if not teacher_rows:
            continue

        # Sort teacher subject rows alphabetically
        teacher_rows = sorted(teacher_rows, key=lambda x: x["nama_mapel"])

        # Calculate total JP on PAGI and SIANG shifts
        total_p = 0
        total_s = 0
        for tr in teacher_rows:
            for i, val in enumerate(tr["jp_per_kelas"]):
                if val == "":
                    continue
                if sorted_classes[i]["shift_operasional"] == "PAGI":
                    total_p += val
                else:
                    total_s += val
        total_jm = total_p + total_s

        guru_row_start = row_idx
        jabatan = "Staf Pengajar"

        for gi, tr in enumerate(teacher_rows):
            row_data = [
                guru_no if gi == 0 else "",
                teacher["nama_guru"] if gi == 0 else "",
                teacher["kode_guru"] if gi == 0 else "",
                jabatan if gi == 0 else "",
                tr["nama_mapel"]
            ]
            row_data.extend(tr["jp_per_kelas"])
            
            # P, S, JM values only on the first row
            row_data.extend([
                total_p if (gi == 0 and total_p > 0) else "",
                total_s if (gi == 0 and total_s > 0) else "",
                total_jm if (gi == 0 and total_jm > 0) else ""
            ])
            
            ws_m.append(row_data)
            ws_m.row_dimensions[row_idx].height = 20
            
            # Style range row
            for c in range(1, total_cols_m + 1):
                cell = ws_m.cell(row=row_idx, column=c)
                cell.border = border_cell
                cell.font = font_data_normal
                if c in (1, 3, 4, col_p, col_s, col_jm) or (6 <= c <= 5 + n_classes):
                    cell.alignment = align_center
                elif c in (2, 5):
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            row_idx += 1

        # Merge vertical fields for teacher identity & P, S, JM values
        if len(teacher_rows) > 1:
            guru_row_end = row_idx - 1
            # Merge No, Name, Kode Guru, Position
            for c_idx in (1, 2, 3, 4):
                style_range(ws_m, start_row=guru_row_start, start_col=c_idx, end_row=guru_row_end, end_col=c_idx, 
                            border=border_cell, alignment=align_center if c_idx != 2 else Alignment(horizontal="left", vertical="center"))
                ws_m.merge_cells(start_row=guru_row_start, start_column=c_idx, end_row=guru_row_end, end_column=c_idx)
            # Merge P, S, JM
            for c_idx in (col_p, col_s, col_jm):
                style_range(ws_m, start_row=guru_row_start, start_col=c_idx, end_row=guru_row_end, end_col=c_idx, 
                            border=border_cell, alignment=align_center, font=font_data_bold)
                ws_m.merge_cells(start_row=guru_row_start, start_column=c_idx, end_row=guru_row_end, end_column=c_idx)
        else:
            # Set font to bold for totals even if only 1 row
            for c_idx in (col_p, col_s, col_jm):
                ws_m.cell(row=guru_row_start, column=c_idx).font = font_data_bold

        guru_no += 1

    # ── 4. Footer Row (JUMLAH JAM PER MINGGU) ──
    foot_row = ["", "JUMLAH JAM PER MINGGU", "", "", ""]
    tot_p = 0
    tot_s = 0
    
    for cls in sorted_classes:
        val = class_total_alloc.get(cls["id_kelas"], 0)
        foot_row.append(val if val > 0 else "")
        if cls["shift_operasional"] == "PAGI":
            tot_p += val
        else:
            tot_s += val
            
    foot_row.extend([
        tot_p if tot_p > 0 else "",
        tot_s if tot_s > 0 else "",
        (tot_p + tot_s) if (tot_p + tot_s) > 0 else ""
    ])
    ws_m.append(foot_row)
    ws_m.row_dimensions[row_idx].height = 22
    
    # Merge label "JUMLAH JAM PER MINGGU" across Cols B to E
    style_range(ws_m, start_row=row_idx, start_col=2, end_row=row_idx, end_col=5, border=border_cell, fill=fill_header, font=font_header_label, alignment=align_center)
    ws_m.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=5)
    
    # Style remaining cells in the footer
    ws_m.cell(row=row_idx, column=1).border = border_cell
    for c in range(6, total_cols_m + 1):
        cell = ws_m.cell(row=row_idx, column=c)
        cell.border = border_cell
        cell.fill = fill_header
        cell.font = font_header_label
        cell.alignment = align_center

    # Column Widths
    ws_m.column_dimensions['A'].width = 5   # No
    ws_m.column_dimensions['B'].width = 28  # Nama Guru
    ws_m.column_dimensions['C'].width = 12  # Kode Guru
    ws_m.column_dimensions['D'].width = 16  # Jabatan
    ws_m.column_dimensions['E'].width = 26  # Mapel
    
    for i in range(n_classes):
        ws_m.column_dimensions[get_column_letter(6 + i)].width = 5  # Kelas cols
    ws_m.column_dimensions[get_column_letter(col_p)].width = 6      # P
    ws_m.column_dimensions[get_column_letter(col_s)].width = 6      # S
    ws_m.column_dimensions[get_column_letter(col_jm)].width = 7     # JM

    # Return file buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    branch_name = get_branch_name().upper()
    filename = f"Jadwal_SMK_11_MARET_{branch_name}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    
    return buf, filename
